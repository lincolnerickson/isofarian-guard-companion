#!/usr/bin/env python3
"""
Build script for The Isofarian Guard 2E Companion App.
Reads the Excel spreadsheet and generates a self-contained HTML file.
"""

import openpyxl
import json
import os

EXCEL_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                          'TIG_2E_Unofficial_Index_Companion_v2.xlsx')
OUTPUT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           'isofarian_companion.html')


def cell_str(val):
    """Convert cell value to clean string."""
    if val is None:
        return ''
    return str(val).strip()


def parse_enemies(wb):
    """Parse Common Bestiary sheet."""
    ws = wb['Common Bestiary']
    enemies = []
    current_name = None

    # Column mapping from row 5:
    # B=Name, C=Rating, D=Att, E=Def, F=AP, G=HP
    # H=Ch1, I=Ch2, J=Ch3, K=Ch4
    # L=Lux, M=Sil, N=Item, O=Speaking Stone
    # P=Metal Frag (★), Q=Bone Frag, R=Feathers, S=Wolf Pelt
    # T=Rough Leather (★★), U=Animal Hide, V=Claw
    # W=Bear Pelt (★★★), X=Horn, Y=Spines
    # Z=Scales (★★★★), AA=Carapace
    # AB=Tenebris Shards (★), AC=Tenebris Skull (★★), AD=Tenebris Essence (★★★)

    drop_cols = {
        'Metal Fragments': 15,   # P (col index 15, 0-based)
        'Bone Fragments': 16,
        'Feathers': 17,
        'Wolf Pelt': 18,
        'Rough Leather': 19,
        'Animal Hide': 20,
        'Claw': 21,
        'Bear Pelt': 22,
        'Horn': 23,
        'Spines': 24,
        'Scales': 25,
        'Carapace': 26,
        'Tenebris Shards': 27,
        'Tenebris Skull': 28,
        'Tenebris Essence': 29,
    }

    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True):
        vals = [cell_str(c) for c in row]
        # Skip empty rows
        if not any(vals):
            continue

        name = vals[1]
        if name:
            current_name = name

        rating = vals[2]
        if not rating:
            continue

        drops = []
        for mat_name, col_idx in drop_cols.items():
            v = vals[col_idx] if col_idx < len(vals) else ''
            if v and v.upper() == 'X':
                drops.append(mat_name)
            elif v and v not in ('', 'X'):
                # Some enemies have location-specific drops
                drops.append(f"{mat_name} ({v})")

        locations = {}
        for ch, idx in [('Chapter 1', 7), ('Chapter 2', 8), ('Chapter 3', 9), ('Chapter 4', 10)]:
            loc = vals[idx] if idx < len(vals) else ''
            if loc:
                locations[ch] = loc

        entry = {
            'name': current_name,
            'rating': rating,
            'attack': vals[3],
            'defense': vals[4],
            'ap': vals[5],
            'hp': vals[6],
            'locations': locations,
            'lux': vals[11],
            'silver': vals[12],
            'itemDrop': vals[13],
            'speakingStoneDrop': vals[14],
            'materialDrops': drops,
        }
        enemies.append(entry)

    return enemies


def parse_armor_weapons(wb):
    """Parse Armor-Weapon Guide sheet."""
    ws = wb['Armor-Weapon Guide']
    items = []

    # Material columns mapping (paired: Qty, 2R)
    mat_cols = [
        ('Metal Fragments', 22, 23),
        ('Bone Fragments', 24, 25),
        ('Feathers', 26, 27),
        ('Wolf Pelt', 28, 29),
        ('Rough Leather', 30, 31),
        ('Animal Hide', 32, 33),
        ('Claw', 34, 35),
        ('Bear Pelt', 36, 37),
        ('Horn', 38, 39),
        ('Spines', 40, 41),
        ('Scales', 42, 43),
        ('Carapace', 44, 45),
        ('Tenebris Shards', 46, 47),
        ('Tenebris Skull', 48, 49),
        ('Tenebris Essence', 50, 51),
    ]

    wood_cols = [
        ('Pine', 53, 54),
        ('Rosewood', 55, 56),
        ('Ash', 57, 58),
        ('Autumn Blaze', 59, 60),
        ('Dogwood', 61, 62),
        ('Cedar', 63, 64),
        ('Cherry', 65, 66),
        ('Ancient Oak', 67, 68),
    ]

    ore_cols = [
        ('Iron', 70, 71),
        ('Silver', 72, 73),
        ('Gold', 74, 75),
        ('Agate', 76, 77),
        ('Crystal', 78, 79),
        ('Diamond', 80, 81),
    ]

    for row in ws.iter_rows(min_row=7, max_row=ws.max_row, values_only=True):
        vals = [cell_str(c) for c in row]
        name = vals[1]
        if not name:
            continue

        materials = {}
        for mat_name, qty_col, rep_col in mat_cols:
            qty = vals[qty_col] if qty_col < len(vals) else ''
            rep = vals[rep_col] if rep_col < len(vals) else ''
            if qty:
                materials[mat_name] = {'qty': qty, 'rep2': rep}

        wood = {}
        for w_name, qty_col, rep_col in wood_cols:
            qty = vals[qty_col] if qty_col < len(vals) else ''
            rep = vals[rep_col] if rep_col < len(vals) else ''
            if qty:
                wood[w_name] = {'qty': qty, 'rep2': rep}

        ores = {}
        for o_name, qty_col, rep_col in ore_cols:
            qty = vals[qty_col] if qty_col < len(vals) else ''
            rep = vals[rep_col] if rep_col < len(vals) else ''
            if qty:
                ores[o_name] = {'qty': qty, 'rep2': rep}

        speaking_stone = vals[19] if 19 < len(vals) else ''
        speaking_stone_rep = vals[20] if 20 < len(vals) else ''

        entry = {
            'name': name,
            'city': vals[2],
            'limitedTo': vals[3],
            'rating': vals[4],
            'type': vals[5],
            'statIncrease': vals[6],
            'stoneSlots': vals[7],
            'bonusChip': vals[8],
            'craftCost': vals[10],
            'sellPrice': vals[11],
            'luxCost': vals[13],
            'prerequisite': vals[15],
            'itemRequired': vals[17],
            'speakingStone': speaking_stone,
            'speakingStoneRep': speaking_stone_rep,
            'materials': materials,
            'wood': wood,
            'ores': ores,
        }
        items.append(entry)

    return items


def parse_accessories(wb):
    """Parse Accessory-Item Guide sheet."""
    ws = wb['Accessory-Item Guide']
    items = []

    mat_cols = [
        ('Metal Fragments', 18, 19),
        ('Bone Fragments', 20, 21),
        ('Feathers', 22, 23),
        ('Wolf Pelt', 24, 25),
        ('Rough Leather', 26, 27),
        ('Animal Hide', 28, 29),
        ('Claw', 30, 31),
        ('Bear Pelt', 32, 33),
        ('Horn', 34, 35),
        ('Spines', 36, 37),
        ('Scales', 38, 39),
        ('Carapace', 40, 41),
        ('Tenebris Shards', 42, 43),
        ('Tenebris Skull', 44, 45),
        ('Tenebris Essence', 46, 47),
    ]

    wood_cols = [
        ('Pine', 49, 50),
        ('Rosewood', 51, 52),
        ('Ash', 53, 54),
        ('Autumn Blaze', 55, 56),
        ('Dogwood', 57, 58),
        ('Cedar', 59, 60),
        ('Cherry', 61, 62),
        ('Ancient Oak', 63, 64),
    ]

    ore_cols = [
        ('Iron', 66, 67),
        ('Silver', 68, 69),
        ('Gold', 70, 71),
        ('Agate', 72, 73),
        ('Crystal', 74, 75),
        ('Diamond', 76, 77),
    ]

    for row in ws.iter_rows(min_row=7, max_row=ws.max_row, values_only=True):
        vals = [cell_str(c) for c in row]
        name = vals[1]
        if not name:
            continue

        materials = {}
        for mat_name, qty_col, rep_col in mat_cols:
            qty = vals[qty_col] if qty_col < len(vals) else ''
            rep = vals[rep_col] if rep_col < len(vals) else ''
            if qty:
                materials[mat_name] = {'qty': qty, 'rep2': rep}

        wood = {}
        for w_name, qty_col, rep_col in wood_cols:
            qty = vals[qty_col] if qty_col < len(vals) else ''
            rep = vals[rep_col] if rep_col < len(vals) else ''
            if qty:
                wood[w_name] = {'qty': qty, 'rep2': rep}

        ores = {}
        for o_name, qty_col, rep_col in ore_cols:
            qty = vals[qty_col] if qty_col < len(vals) else ''
            rep = vals[rep_col] if rep_col < len(vals) else ''
            if qty:
                ores[o_name] = {'qty': qty, 'rep2': rep}

        speaking_stone = vals[16] if 16 < len(vals) else ''
        speaking_stone_rep = vals[17] if 17 < len(vals) else ''

        entry = {
            'name': name,
            'city': vals[2],
            'type': vals[3],
            'usableInField': vals[4],
            'effect': vals[5],
            'craftCost': vals[7],
            'sellPrice': vals[8],
            'luxCost': vals[10],
            'prerequisite': vals[12],
            'itemRequired': vals[14],
            'speakingStone': speaking_stone,
            'speakingStoneRep': speaking_stone_rep,
            'materials': materials,
            'wood': wood,
            'ores': ores,
        }
        items.append(entry)

    return items


def parse_market(wb):
    """Parse Market Guide sheet."""
    ws = wb['Market Guide']
    items = []

    towns = ['Mir', 'Razdor', 'Ryba', 'Silny', 'Strofa', 'Vouno', 'Fort Istra Apothecary']
    # Each town has 3 cols: Buy, Buy 2Rep, Sell
    # Starting at col D (index 3)

    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True):
        vals = [cell_str(c) for c in row]
        name = vals[1]
        if not name:
            continue

        effect = vals[2]
        prices = {}
        col = 3
        for town in towns:
            buy = vals[col] if col < len(vals) else ''
            buy2r = vals[col + 1] if col + 1 < len(vals) else ''
            sell = vals[col + 2] if col + 2 < len(vals) else ''
            if buy or buy2r or sell:
                prices[town] = {'buy': buy, 'buy2Rep': buy2r, 'sell': sell}
            col += 3

        entry = {
            'name': name,
            'effect': effect,
            'prices': prices,
        }
        items.append(entry)

    return items


def parse_buildings(wb):
    """Parse Ft. Istra Buildings sheet."""
    ws = wb['Ft. Istra Buildings']
    buildings = []

    wood_cols = [('Pine', 4), ('Rosewood', 5), ('Ash', 6), ('Autumn Blaze', 7),
                 ('Dogwood', 8), ('Cedar', 9), ('Cherry', 10), ('Ancient Oak', 11)]
    ore_cols = [('Iron', 13), ('Silver', 14), ('Gold', 15),
                ('Agate', 16), ('Crystal', 17), ('Diamond', 18)]

    for row in ws.iter_rows(min_row=6, max_row=20, values_only=True):
        vals = [cell_str(c) for c in row]
        name = vals[1]
        if not name:
            continue

        item_req = vals[2]
        wood = {}
        for w_name, idx in wood_cols:
            v = vals[idx] if idx < len(vals) else ''
            if v:
                wood[w_name] = v

        ores = {}
        for o_name, idx in ore_cols:
            v = vals[idx] if idx < len(vals) else ''
            if v:
                ores[o_name] = v

        buildings.append({
            'name': name,
            'itemRequired': item_req,
            'wood': wood,
            'ores': ores,
        })

    # Parse harvesting locations (row 21)
    harvest = {}
    for row in ws.iter_rows(min_row=21, max_row=21, values_only=True):
        vals = [cell_str(c) for c in row]
        for w_name, idx in wood_cols:
            v = vals[idx] if idx < len(vals) else ''
            if v:
                harvest[w_name] = v
        for o_name, idx in ore_cols:
            v = vals[idx] if idx < len(vals) else ''
            if v:
                harvest[o_name] = v

    # Parse Lux costs (row 23)
    lux_costs = {}
    for row in ws.iter_rows(min_row=23, max_row=23, values_only=True):
        vals = [cell_str(c) for c in row]
        for w_name, idx in wood_cols:
            v = vals[idx] if idx < len(vals) else ''
            if v:
                lux_costs[w_name] = v
        for o_name, idx in ore_cols:
            v = vals[idx] if idx < len(vals) else ''
            if v:
                lux_costs[o_name] = v

    return buildings, harvest, lux_costs


def parse_speaking_stones(wb):
    """Parse Speaking Stone Bonuses sheet."""
    ws = wb['Speaking Stone Bonuses']
    stones = []

    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=True):
        vals = [cell_str(c) for c in row]
        name = vals[1]
        if not name:
            continue

        stones.append({
            'name': name,
            'available': vals[2],
            'element': vals[3],
            'color': vals[4],
            'bonus': vals[5],
            'lapidaryExchange': vals[6],
        })

    return stones


def parse_prereqs(wb):
    """Parse Blacksmith Pre-req Guide for upgrade chains."""
    ws = wb['Blacksmith Pre-req Guide']
    chains = []
    current_chain = []

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        vals = [cell_str(c) for c in row]
        name = vals[1]

        if not name:
            if current_chain:
                chains.append(current_chain)
                current_chain = []
            continue

        current_chain.append({
            'name': name,
            'city': vals[2],
            'limitedTo': vals[3],
            'rating': vals[4],
            'type': vals[5],
            'statIncrease': vals[6],
            'stoneSlots': vals[7],
            'bonusChip': vals[8],
        })

    if current_chain:
        chains.append(current_chain)

    return chains


# ---------------------------------------------------------------------------
# Map graph data: node positions (px on 3000x4511 map) + edges
# Positions are best-effort traces from map-of-isofar.png.
# Use the in-app editor mode to refine positions interactively.
# ---------------------------------------------------------------------------
MAP_GRAPH = {
    "nodes": {},
    "edges": []
}

# Lookup table for location string -> node ID mapping
SPECIAL_AREA_NAME_TO_ID = {
    "FW - Ice Fields": "fw_ice_fields",
    "FW - Mount Nebesa": "fw_mount_nebesa",
    "FW - Reka Glacier": "fw_reka_glacier",
    "FW - Room of Columns": "fw_room_of_columns",
    "FW - Skryvat Temple": "fw_skryvat_temple",
    "FW - The Broken Lands": "fw_broken_lands",
    "FW - Uchitel Span": "fw_uchitel_span",
    "FW - Urok Span": "fw_urok_span",
    "FW - Vniz Path": "fw_vniz_path",
    "IC - Abandoned Quarters": "ic_abandoned_quarters",
    "IC - Abandoned Quartes": "ic_abandoned_quarters",  # typo in data
    "IC - Frozen Lake": "ic_frozen_lake",
    "IC - Glacial Worm Bones": "ic_glacial_worm_bones",
    "IC - Hall of Ice": "ic_hall_of_ice",
    "IC - Old Armory": "ic_old_armory",
    "IC - Ossuary": "ic_ossuary",
}


def enrich_map_graph(data):
    """Add chapter, enemy, and resource metadata to map graph nodes."""
    import copy, re
    graph = copy.deepcopy(MAP_GRAPH)
    nodes = graph["nodes"]

    # Initialize metadata on each node
    for nid, node in nodes.items():
        node.setdefault("chapters", [])
        node.setdefault("enemies", [])
        node.setdefault("resources", [])

    def parse_location_string(loc_str):
        """Parse a location string like '1, 5, FW - Ice Fields' into node IDs."""
        node_ids = []
        parts = [p.strip().rstrip('.') for p in loc_str.split(',')]
        i = 0
        while i < len(parts):
            p = parts[i].strip()
            if not p:
                i += 1
                continue
            # Check for multi-word special area names
            # Special areas start with "FW" or "IC"
            if p.startswith('FW') or p.startswith('IC'):
                # Reconstruct the full name (may have been split by commas in the name)
                # Actually the names don't have commas, but they're comma-separated in the list
                full_name = p
                nid = SPECIAL_AREA_NAME_TO_ID.get(full_name)
                if nid:
                    node_ids.append(nid)
            else:
                try:
                    num = int(p)
                    node_ids.append(str(num))
                except ValueError:
                    pass
            i += 1
        return node_ids

    # Process enemies to add chapter and enemy info to nodes
    for enemy in data["enemies"]:
        for chapter, loc_str in enemy["locations"].items():
            ch_num = chapter.replace("Chapter ", "")
            nids = parse_location_string(loc_str)
            for nid in nids:
                if nid in nodes:
                    if ch_num not in nodes[nid]["chapters"]:
                        nodes[nid]["chapters"].append(ch_num)
                    if enemy["name"] not in nodes[nid]["enemies"]:
                        nodes[nid]["enemies"].append(enemy["name"])

    # Process harvest locations to add resource info
    for resource, loc_str in data["harvestLocations"].items():
        parts = [p.strip() for p in str(loc_str).split(',')]
        for p in parts:
            try:
                nid = str(int(p))
                if nid in nodes:
                    if resource not in nodes[nid]["resources"]:
                        nodes[nid]["resources"].append(resource)
            except ValueError:
                pass

    # Sort chapters on each node
    for nid, node in nodes.items():
        node["chapters"] = sorted(set(node["chapters"]))

    return graph


def build_html(data):
    """Generate the self-contained HTML file."""
    data_json = json.dumps(data, ensure_ascii=False, indent=None)
    # Prevent </script> injection and HTML entity issues in inline script
    data_json = data_json.replace('<', '\\u003c')
    data_json = data_json.replace('>', '\\u003e')
    data_json = data_json.replace('&', '\\u0026')

    html = r'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<meta http-equiv="Content-Security-Policy" content="default-src 'none'; script-src 'unsafe-inline'; style-src 'unsafe-inline'; img-src 'self' data: blob:;">
<title>The Isofarian Guard 2E - Companion</title>
<style>
:root {
  --bg: #1a1a2e;
  --bg2: #16213e;
  --bg3: #0f3460;
  --accent: #e94560;
  --gold: #f0c040;
  --silver: #c0c0c0;
  --text: #e0e0e0;
  --text2: #a0a0b0;
  --link: #64b5f6;
  --green: #66bb6a;
  --red: #ef5350;
  --orange: #ffa726;
  --purple: #ab47bc;
  --star: #f0c040;
}
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: 'Segoe UI', system-ui, sans-serif; background: var(--bg); color: var(--text); min-height: 100vh; }
a { color: var(--link); cursor: pointer; text-decoration: none; }
a:hover { text-decoration: underline; color: #90caf9; }

.header { background: linear-gradient(135deg, var(--bg2), var(--bg3)); padding: 16px 24px; border-bottom: 2px solid var(--accent); display: flex; align-items: center; gap: 16px; flex-wrap: wrap; }
.header h1 { font-size: 1.4rem; color: var(--gold); white-space: nowrap; }
.search-box { flex: 1; min-width: 200px; max-width: 400px; }
.search-box input { width: 100%; padding: 8px 12px; border-radius: 6px; border: 1px solid var(--bg3); background: var(--bg); color: var(--text); font-size: 0.95rem; }
.search-box input:focus { outline: none; border-color: var(--accent); }

.tabs { display: flex; background: var(--bg2); border-bottom: 1px solid #333; overflow-x: auto; }
.tab { padding: 10px 20px; cursor: pointer; color: var(--text2); border-bottom: 3px solid transparent; white-space: nowrap; font-size: 0.9rem; transition: all 0.2s; }
.tab:hover { color: var(--text); background: rgba(255,255,255,0.05); }
.tab.active { color: var(--gold); border-bottom-color: var(--accent); }

.content { padding: 16px; max-width: 1400px; margin: 0 auto; }
.panel { display: none; }
.panel.active { display: block; }

/* Cards */
.card-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(340px, 1fr)); gap: 12px; }
.card { background: var(--bg2); border: 1px solid #333; border-radius: 8px; padding: 14px; transition: border-color 0.2s; }
.card:hover { border-color: var(--accent); }
.card-title { font-size: 1.1rem; font-weight: 700; color: var(--gold); margin-bottom: 6px; display: flex; align-items: center; gap: 8px; }
.card-subtitle { font-size: 0.85rem; color: var(--text2); margin-bottom: 8px; }
.stat-row { display: flex; flex-wrap: wrap; gap: 6px; margin-bottom: 6px; }
.stat { background: var(--bg); padding: 3px 8px; border-radius: 4px; font-size: 0.82rem; display: inline-flex; align-items: center; gap: 4px; }
.stat-label { color: var(--text2); }
.stat-value { color: var(--text); font-weight: 600; }
.stat.attack .stat-value { color: var(--red); }
.stat.defense .stat-value { color: #42a5f5; }
.stat.hp .stat-value { color: var(--green); }
.stat.ap .stat-value { color: var(--orange); }
.stat.lux .stat-value { color: var(--gold); }
.stat.silver .stat-value { color: var(--silver); }

.rating { color: var(--star); letter-spacing: -2px; font-size: 1rem; }

.section-label { font-size: 0.78rem; color: var(--accent); text-transform: uppercase; font-weight: 600; margin-top: 8px; margin-bottom: 4px; letter-spacing: 0.5px; }
.tag-list { display: flex; flex-wrap: wrap; gap: 4px; }
.tag { background: rgba(100,181,246,0.15); color: var(--link); padding: 2px 8px; border-radius: 3px; font-size: 0.8rem; cursor: pointer; border: 1px solid transparent; }
.tag:hover { border-color: var(--link); background: rgba(100,181,246,0.25); }
.tag.location { background: rgba(102,187,106,0.15); color: var(--green); }
.tag.location:hover { border-color: var(--green); background: rgba(102,187,106,0.25); }
.tag.enemy { background: rgba(239,83,80,0.15); color: var(--red); }
.tag.enemy:hover { border-color: var(--red); background: rgba(239,83,80,0.25); }
.tag.item { background: rgba(171,71,188,0.15); color: var(--purple); }
.tag.item:hover { border-color: var(--purple); background: rgba(171,71,188,0.25); }
.tag.craft { background: rgba(240,192,64,0.15); color: var(--gold); }
.tag.craft:hover { border-color: var(--gold); background: rgba(240,192,64,0.25); }

.locations-list { font-size: 0.82rem; color: var(--text2); margin-top: 4px; }
.locations-list span { color: var(--green); }

/* Detail overlay */
.overlay { display: none; position: fixed; top: 0; left: 0; right: 0; bottom: 0; background: rgba(0,0,0,0.7); z-index: 100; overflow-y: auto; padding: 40px 16px; }
.overlay.active { display: flex; justify-content: center; align-items: flex-start; }
.detail-panel { background: var(--bg2); border: 2px solid var(--accent); border-radius: 12px; padding: 24px; max-width: 700px; width: 100%; position: relative; }
.detail-panel .close-btn { position: absolute; top: 12px; right: 16px; cursor: pointer; font-size: 1.5rem; color: var(--text2); background: none; border: none; }
.detail-panel .close-btn:hover { color: var(--accent); }
.detail-panel h2 { color: var(--gold); margin-bottom: 4px; }
.detail-panel h3 { color: var(--accent); font-size: 0.9rem; margin-top: 14px; margin-bottom: 6px; text-transform: uppercase; }
.detail-section { margin-bottom: 10px; }

.recipe-table { width: 100%; border-collapse: collapse; margin-top: 6px; }
.recipe-table th, .recipe-table td { padding: 6px 10px; text-align: left; border-bottom: 1px solid #333; font-size: 0.85rem; }
.recipe-table th { color: var(--text2); font-weight: 600; }

.prereq-chain { display: flex; align-items: center; gap: 6px; flex-wrap: wrap; margin: 6px 0; }
.prereq-chain .chain-item { background: var(--bg); padding: 4px 10px; border-radius: 4px; font-size: 0.85rem; cursor: pointer; color: var(--link); }
.prereq-chain .chain-item:hover { background: rgba(100,181,246,0.15); }
.prereq-chain .chain-arrow { color: var(--text2); }

/* Filter controls */
.filters { display: flex; flex-wrap: wrap; gap: 8px; margin-bottom: 14px; align-items: center; }
.filter-btn { padding: 5px 14px; border-radius: 20px; border: 1px solid #444; background: var(--bg); color: var(--text2); cursor: pointer; font-size: 0.82rem; transition: all 0.2s; }
.filter-btn:hover, .filter-btn.active { border-color: var(--accent); color: var(--text); background: rgba(233,69,96,0.15); }

/* Table view */
.data-table { width: 100%; border-collapse: collapse; margin-top: 8px; }
.data-table th { background: var(--bg3); padding: 8px 10px; text-align: left; font-size: 0.82rem; color: var(--text2); position: sticky; top: 0; }
.data-table td { padding: 6px 10px; border-bottom: 1px solid #2a2a3e; font-size: 0.85rem; }
.data-table tr:hover td { background: rgba(255,255,255,0.03); }

.empty-msg { text-align: center; color: var(--text2); padding: 40px; font-size: 1rem; }

/* Market table */
.market-table { width: 100%; border-collapse: collapse; font-size: 0.82rem; overflow-x: auto; display: block; }
.market-table th, .market-table td { padding: 6px 8px; border: 1px solid #333; white-space: nowrap; }
.market-table th { background: var(--bg3); color: var(--text2); }
.market-table .town-header { background: var(--bg3); color: var(--gold); text-align: center; }
.market-table .buy { color: var(--red); }
.market-table .sell { color: var(--green); }

/* Route Planner */
.rp-container { display: flex; gap: 0; height: calc(100vh - 120px); margin: -16px; }
.rp-sidebar { width: 320px; min-width: 280px; background: var(--bg2); border-right: 1px solid #333; padding: 16px; overflow-y: auto; flex-shrink: 0; }
.rp-map-area { flex: 1; position: relative; overflow: hidden; background: #111; }
.rp-map-area canvas { display: block; cursor: grab; }
.rp-map-area canvas:active { cursor: grabbing; }
.rp-map-area canvas.editor-mode { cursor: crosshair; }

.rp-sidebar h3 { color: var(--gold); font-size: 0.95rem; margin-bottom: 10px; }
.rp-sidebar label { display: block; font-size: 0.82rem; color: var(--text2); margin: 10px 0 4px; }
.rp-sidebar select, .rp-sidebar input[type="text"] {
  width: 100%; padding: 7px 10px; border-radius: 5px; border: 1px solid #444;
  background: var(--bg); color: var(--text); font-size: 0.88rem;
}
.rp-sidebar select:focus, .rp-sidebar input[type="text"]:focus { border-color: var(--accent); outline: none; }
.rp-btn { display: inline-block; padding: 8px 18px; border-radius: 6px; border: none; cursor: pointer;
  font-size: 0.88rem; font-weight: 600; transition: all 0.2s; margin-top: 10px; }
.rp-btn-primary { background: var(--accent); color: #fff; }
.rp-btn-primary:hover { background: #d63050; }
.rp-btn-secondary { background: var(--bg3); color: var(--text); border: 1px solid #444; }
.rp-btn-secondary:hover { background: #1a3a6e; }
.rp-btn-sm { padding: 4px 10px; font-size: 0.78rem; margin-top: 0; }

.rp-route-results { margin-top: 14px; }
.rp-step { display: flex; gap: 10px; padding: 8px; background: var(--bg); border-radius: 6px;
  margin-bottom: 6px; font-size: 0.85rem; align-items: flex-start; cursor: pointer; border: 1px solid transparent; }
.rp-step:hover { border-color: var(--accent); }
.rp-step-num { background: var(--accent); color: #fff; width: 22px; height: 22px; border-radius: 50%;
  display: flex; align-items: center; justify-content: center; font-size: 0.75rem; font-weight: 700; flex-shrink: 0; }
.rp-step-info { flex: 1; }
.rp-step-node { font-weight: 600; color: var(--gold); }
.rp-step-mats { color: var(--text2); font-size: 0.8rem; margin-top: 2px; }
.rp-mat-link { color: var(--link); cursor: pointer; text-decoration: underline; text-decoration-style: dotted; }
.rp-mat-link:hover { color: #90caf9; }
.rp-mat-detail { margin-top: 4px; padding: 6px 8px; background: var(--bg2); border-left: 2px solid var(--accent); border-radius: 0 4px 4px 0; font-size: 0.78rem; }
.rp-mat-detail .rp-md-enemy { color: var(--red); font-weight: 600; }
.rp-mat-detail .rp-md-source { color: var(--green); }
.rp-mat-detail .rp-md-market { color: var(--gold); }
.rp-step-dist { color: var(--text2); font-size: 0.75rem; }

.rp-tooltip { position: absolute; background: var(--bg2); border: 1px solid var(--accent); border-radius: 6px;
  padding: 8px 12px; font-size: 0.82rem; pointer-events: none; z-index: 50; max-width: 250px;
  box-shadow: 0 4px 12px rgba(0,0,0,0.5); display: none; }
.rp-tooltip .tt-name { font-weight: 700; color: var(--gold); }
.rp-tooltip .tt-info { color: var(--text2); margin-top: 4px; }

.rp-editor-bar { display: flex; gap: 6px; align-items: center; margin-top: 12px; padding-top: 10px; border-top: 1px solid #333; flex-wrap: wrap; }
.rp-search-filter { margin-top: 6px; }

.rp-summary { background: var(--bg); border-radius: 6px; padding: 10px; margin-top: 10px; font-size: 0.85rem; }
.rp-summary .label { color: var(--text2); }
.rp-summary .value { color: var(--gold); font-weight: 600; }

.rp-legend { display: flex; gap: 12px; flex-wrap: wrap; margin-top: 8px; font-size: 0.78rem; color: var(--text2); }
.rp-legend-dot { width: 10px; height: 10px; border-radius: 50%; display: inline-block; margin-right: 4px; vertical-align: middle; }

/* Responsive */
@media (max-width: 600px) {
  .card-grid { grid-template-columns: 1fr; }
  .header { padding: 12px; }
  .header h1 { font-size: 1.1rem; }
  .content { padding: 10px; }
  .tabs { font-size: 0.85rem; }
  .tab { padding: 8px 14px; }
  .rp-container { flex-direction: column; height: auto; }
  .rp-sidebar { width: 100%; border-right: none; border-bottom: 1px solid #333; }
  .rp-map-area { height: 60vh; }
}
</style>
</head>
<body>

<div class="header">
  <h1>The Isofarian Guard 2E Companion</h1>
  <div class="search-box">
    <input type="text" id="globalSearch" placeholder="Search enemies, items, materials, locations..." />
  </div>
</div>

<div class="tabs" id="tabs">
  <div class="tab active" data-tab="enemies">Enemies</div>
  <div class="tab" data-tab="armor-weapons">Armor & Weapons</div>
  <div class="tab" data-tab="accessories">Accessories & Items</div>
  <div class="tab" data-tab="market">Market</div>
  <div class="tab" data-tab="buildings">Ft. Istra Buildings</div>
  <div class="tab" data-tab="stones">Speaking Stones</div>
  <div class="tab" data-tab="materials">Material Finder</div>
  <div class="tab" data-tab="route-planner">Route Planner</div>
</div>

<div class="content" id="content">
  <!-- Enemies -->
  <div class="panel active" id="panel-enemies">
    <div class="filters" id="enemy-filters">
      <span style="color:var(--text2);font-size:0.85rem;">Rating:</span>
      <button class="filter-btn active" data-filter="all">All</button>
      <button class="filter-btn" data-filter="★">★</button>
      <button class="filter-btn" data-filter="★★">★★</button>
      <button class="filter-btn" data-filter="★★★">★★★</button>
      <button class="filter-btn" data-filter="★★★★">★★★★</button>
    </div>
    <div class="card-grid" id="enemy-grid"></div>
  </div>

  <!-- Armor & Weapons -->
  <div class="panel" id="panel-armor-weapons">
    <div class="filters" id="aw-filters">
      <span style="color:var(--text2);font-size:0.85rem;">Type:</span>
      <button class="filter-btn active" data-filter="all">All</button>
      <button class="filter-btn" data-filter="Armor">Armor</button>
      <button class="filter-btn" data-filter="Weapon">Weapon</button>
      <span style="color:var(--text2);font-size:0.85rem;margin-left:10px;">Rating:</span>
      <button class="filter-btn active" data-rating="all">All</button>
      <button class="filter-btn" data-rating="★">★</button>
      <button class="filter-btn" data-rating="★★">★★</button>
      <button class="filter-btn" data-rating="★★★">★★★</button>
      <button class="filter-btn" data-rating="★★★★">★★★★</button>
    </div>
    <div class="card-grid" id="aw-grid"></div>
  </div>

  <!-- Accessories -->
  <div class="panel" id="panel-accessories">
    <div class="filters" id="acc-filters">
      <span style="color:var(--text2);font-size:0.85rem;">Type:</span>
      <button class="filter-btn active" data-filter="all">All</button>
      <button class="filter-btn" data-filter="Accessory">Accessory</button>
      <button class="filter-btn" data-filter="Item">Item</button>
    </div>
    <div class="card-grid" id="acc-grid"></div>
  </div>

  <!-- Market -->
  <div class="panel" id="panel-market">
    <div id="market-content"></div>
  </div>

  <!-- Buildings -->
  <div class="panel" id="panel-buildings">
    <div id="buildings-content"></div>
  </div>

  <!-- Speaking Stones -->
  <div class="panel" id="panel-stones">
    <div class="card-grid" id="stones-grid"></div>
  </div>

  <!-- Material Finder -->
  <div class="panel" id="panel-materials">
    <p style="color:var(--text2);margin-bottom:12px;">Click any material to see which enemies drop it, where to buy it, and what it's used to craft.</p>
    <div class="tag-list" id="material-index" style="gap:6px;"></div>
    <div id="material-detail" style="margin-top:16px;"></div>
  </div>

  <!-- Route Planner -->
  <div class="panel" id="panel-route-planner">
    <div class="rp-container">
      <div class="rp-sidebar">
        <h3>Route Planner</h3>
        <label for="rp-item-select">Item to Craft</label>
        <input type="text" id="rp-item-search" placeholder="Search items..." class="rp-search-filter" />
        <select id="rp-item-select"><option value="">-- Select an item --</option></select>
        <label for="rp-start-select">Starting Location</label>
        <select id="rp-start-select"></select>
        <label for="rp-chapter-select">Current Chapter</label>
        <select id="rp-chapter-select">
          <option value="1">Chapter 1</option>
          <option value="2">Chapter 2</option>
          <option value="3">Chapter 3</option>
          <option value="4">Chapter 4</option>
        </select>
        <button class="rp-btn rp-btn-primary" id="rp-calculate" style="width:100%;">Calculate Route</button>
        <div id="rp-route-summary" class="rp-summary" style="display:none;"></div>
        <div id="rp-route-results" class="rp-route-results"></div>
        <div class="rp-editor-bar">
          <button class="rp-btn rp-btn-secondary rp-btn-sm" id="rp-edit-toggle">Edit Map</button>
          <button class="rp-btn rp-btn-secondary rp-btn-sm" id="rp-edit-connect" style="display:none;">Connect Nodes</button>
          <button class="rp-btn rp-btn-secondary rp-btn-sm" id="rp-edit-save" style="display:none;">Save</button>
          <button class="rp-btn rp-btn-secondary rp-btn-sm" id="rp-edit-reset" style="display:none;">Reset</button>
          <button class="rp-btn rp-btn-secondary rp-btn-sm" id="rp-edit-export" style="display:none;">Export JSON</button>
        </div>
        <div id="rp-editor-help" style="display:none;font-size:0.78rem;color:var(--text2);margin-top:8px;padding:8px;background:var(--bg);border-radius:4px;">
          <b style="color:var(--gold);">Editor Controls:</b><br>
          <b>Click</b> empty space &rarr; add node<br>
          <b>Drag</b> a node &rarr; move it<br>
          <b>Connect Nodes</b> &rarr; click nodes to draw edges between them (click same node or empty space to stop)<br>
          <b>Right-click</b> node &rarr; delete it<br>
          <b>Save</b> stores to browser. <b>Export</b> copies JSON.
        </div>
        <div class="rp-legend">
          <span><span class="rp-legend-dot" style="background:#f0c040;"></span>Town</span>
          <span><span class="rp-legend-dot" style="background:#64b5f6;"></span>Node</span>
          <span><span class="rp-legend-dot" style="background:#ab47bc;"></span>Special</span>
          <span><span class="rp-legend-dot" style="background:#ef5350;"></span>Route Stop</span>
        </div>
      </div>
      <div class="rp-map-area" id="rp-map-area">
        <canvas id="rp-canvas"></canvas>
        <div class="rp-tooltip" id="rp-tooltip"></div>
      </div>
    </div>
  </div>
</div>

<div class="overlay" id="overlay">
  <div class="detail-panel" id="detail-panel">
    <button class="close-btn" id="close-detail">&times;</button>
    <div id="detail-content"></div>
  </div>
</div>

<script>
const DATA = ''' + data_json + r''';

// --- Utility ---
function esc(s) { if (!s) return ''; return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;'); }
function escJs(s) { if (!s) return ''; return String(s).replace(/\\/g,'\\\\').replace(/'/g,"\\'").replace(/"/g,'\\"').replace(/\n/g,'\\n'); }

function formatCost(s) {
  if (!s || s === '-') return s || '';
  // Split on newlines, then fix each part: "20RAZDOR/SILNY" -> "$20 Razdor/Silny"
  return s.split(/\n/).map(part => {
    part = part.trim();
    if (!part) return '';
    // Insert space between number and letters: "20RAZDOR" -> "20 RAZDOR"
    part = part.replace(/^(\d+)([A-Za-z])/, '$1 $2');
    // Add $ before leading number
    part = part.replace(/^(\d)/, '$$$1');
    return part;
  }).filter(Boolean).join(', ');
}

function ratingStars(r) {
  if (!r) return '';
  return '<span class="rating">' + esc(r) + '</span>';
}

// --- Cross-reference index builders ---
const materialToEnemies = {};
const materialToCraft = {};
const materialToMarket = {};

function buildIndices() {
  // Enemies -> materials
  DATA.enemies.forEach(e => {
    e.materialDrops.forEach(d => {
      let matName = d.includes('(') ? d.substring(0, d.indexOf('(')).trim() : d;
      if (!materialToEnemies[matName]) materialToEnemies[matName] = [];
      materialToEnemies[matName].push(e);
    });
  });

  // Armor/Weapons -> materials
  DATA.armorWeapons.forEach(item => {
    [item.materials, item.wood, item.ores].forEach(group => {
      Object.keys(group).forEach(mat => {
        if (!materialToCraft[mat]) materialToCraft[mat] = [];
        materialToCraft[mat].push({ name: item.name, type: 'armor-weapon', item });
      });
    });
    if (item.speakingStone) {
      let sName = item.speakingStone.replace(/\s*x\d+/i, '');
      if (!materialToCraft[sName]) materialToCraft[sName] = [];
      materialToCraft[sName].push({ name: item.name, type: 'armor-weapon', item });
    }
  });

  // Accessories -> materials
  DATA.accessories.forEach(item => {
    [item.materials, item.wood, item.ores].forEach(group => {
      Object.keys(group).forEach(mat => {
        if (!materialToCraft[mat]) materialToCraft[mat] = [];
        materialToCraft[mat].push({ name: item.name, type: 'accessory', item });
      });
    });
    if (item.speakingStone) {
      let sName = item.speakingStone.replace(/\s*x\d+/i, '');
      if (!materialToCraft[sName]) materialToCraft[sName] = [];
      materialToCraft[sName].push({ name: item.name, type: 'accessory', item });
    }
  });

  // Market -> materials
  DATA.market.forEach(item => {
    materialToMarket[item.name] = item;
  });
}

// --- All unique materials ---
function getAllMaterials() {
  const mats = new Set();
  DATA.enemies.forEach(e => e.materialDrops.forEach(d => {
    mats.add(d.includes('(') ? d.substring(0, d.indexOf('(')).trim() : d);
  }));
  DATA.armorWeapons.forEach(i => {
    [i.materials, i.wood, i.ores].forEach(g => Object.keys(g).forEach(m => mats.add(m)));
  });
  DATA.accessories.forEach(i => {
    [i.materials, i.wood, i.ores].forEach(g => Object.keys(g).forEach(m => mats.add(m)));
  });
  return [...mats].sort();
}

// --- Navigation ---
function showDetail(type, identifier) {
  const overlay = document.getElementById('overlay');
  const content = document.getElementById('detail-content');

  if (type === 'enemy') {
    const entries = DATA.enemies.filter(e => e.name === identifier);
    if (!entries.length) return;
    let html = '<h2>' + esc(identifier) + '</h2>';
    html += '<div class="card-subtitle">Enemy</div>';

    entries.forEach(e => {
      html += '<div style="margin-top:12px;padding:10px;background:var(--bg);border-radius:6px;">';
      html += '<div style="display:flex;align-items:center;gap:8px;margin-bottom:6px;">' + ratingStars(e.rating) + '</div>';
      html += '<div class="stat-row">';
      html += '<span class="stat attack"><span class="stat-label">ATK</span><span class="stat-value">' + esc(e.attack) + '</span></span>';
      html += '<span class="stat defense"><span class="stat-label">DEF</span><span class="stat-value">' + esc(e.defense) + '</span></span>';
      html += '<span class="stat ap"><span class="stat-label">AP</span><span class="stat-value">' + esc(e.ap) + '</span></span>';
      html += '<span class="stat hp"><span class="stat-label">HP</span><span class="stat-value">' + esc(e.hp) + '</span></span>';
      if (e.lux) html += '<span class="stat lux"><span class="stat-label">Lux</span><span class="stat-value">' + esc(e.lux) + '</span></span>';
      if (e.silver) html += '<span class="stat silver"><span class="stat-label">Silver</span><span class="stat-value">' + esc(e.silver) + '</span></span>';
      html += '</div>';

      if (Object.keys(e.locations).length) {
        html += '<div class="section-label">Locations</div>';
        Object.entries(e.locations).forEach(([ch, loc]) => {
          html += '<div style="font-size:0.85rem;margin-bottom:2px;"><span style="color:var(--text2);">' + esc(ch) + ':</span> <span style="color:var(--green);">' + esc(loc) + '</span></div>';
        });
      }

      if (e.materialDrops.length) {
        html += '<div class="section-label">Material Drops</div><div class="tag-list">';
        e.materialDrops.forEach(d => {
          let matName = d.includes('(') ? d.substring(0, d.indexOf('(')).trim() : d;
          html += '<a class="tag" onclick="showDetail(\'material\',\'' + escJs(matName) + '\')">' + esc(d) + '</a>';
        });
        html += '</div>';
      }

      if (e.itemDrop) {
        html += '<div class="section-label">Special Drops</div>';
        html += '<span class="tag item">' + esc(e.itemDrop) + '</span>';
      }
      if (e.speakingStoneDrop) {
        html += '<div class="section-label">Speaking Stone Drop</div>';
        html += '<span class="tag item">' + esc(e.speakingStoneDrop) + '</span>';
      }
      html += '</div>';
    });

    content.innerHTML = html;

  } else if (type === 'craft' || type === 'armor-weapon' || type === 'accessory') {
    let item = DATA.armorWeapons.find(i => i.name === identifier);
    let itemType = 'armor-weapon';
    if (!item) {
      item = DATA.accessories.find(i => i.name === identifier);
      itemType = 'accessory';
    }
    if (!item) return;

    let html = '<h2>' + esc(item.name) + '</h2>';
    html += '<div class="card-subtitle">';
    html += esc(item.type || '') + (item.city ? ' &mdash; ' + esc(item.city) : '') + (item.limitedTo ? ' (Limited: ' + esc(item.limitedTo) + ')' : '');
    html += '</div>';

    html += '<div class="stat-row">';
    if (item.rating) html += '<span class="stat">' + ratingStars(item.rating) + '</span>';
    if (item.statIncrease) html += '<span class="stat"><span class="stat-label">Stat</span><span class="stat-value">' + esc(item.statIncrease) + '</span></span>';
    if (item.stoneSlots) html += '<span class="stat"><span class="stat-label">Slots</span><span class="stat-value">' + esc(item.stoneSlots) + '</span></span>';
    if (item.bonusChip) html += '<span class="stat"><span class="stat-label">Bonus</span><span class="stat-value">' + esc(item.bonusChip) + '</span></span>';
    html += '</div>';

    if (item.effect) {
      html += '<div style="font-size:0.85rem;color:var(--text);margin:8px 0;padding:8px;background:var(--bg);border-radius:4px;">' + esc(item.effect) + '</div>';
    }

    // Cost
    html += '<div class="stat-row" style="margin-top:8px;">';
    if (item.craftCost && item.craftCost !== '-') html += '<span class="stat lux"><span class="stat-label">Craft</span><span class="stat-value">' + esc(formatCost(item.craftCost)) + '</span></span>';
    if (item.sellPrice && item.sellPrice !== '-') html += '<span class="stat silver"><span class="stat-label">Sell</span><span class="stat-value">' + esc(formatCost(item.sellPrice)) + '</span></span>';
    if (item.luxCost && item.luxCost !== '-') html += '<span class="stat lux"><span class="stat-label">Lux</span><span class="stat-value">' + esc(formatCost(item.luxCost)) + '</span></span>';
    html += '</div>';

    // Prerequisite
    if (item.prerequisite) {
      html += '<h3>Prerequisite Equipment</h3>';
      html += '<a class="tag craft" onclick="showDetail(\'craft\',\'' + escJs(item.prerequisite) + '\')">' + esc(item.prerequisite) + '</a>';

      // Show chain
      let chain = findPrereqChain(item.name);
      if (chain && chain.length > 1) {
        html += '<div class="prereq-chain">';
        chain.forEach((c, i) => {
          html += '<a class="chain-item" onclick="showDetail(\'craft\',\'' + escJs(c.name) + '\')">' + esc(c.name) + ' ' + ratingStars(c.rating) + '</a>';
          if (i < chain.length - 1) html += '<span class="chain-arrow">&rarr;</span>';
        });
        html += '</div>';
      }
    }

    // Required items
    if (item.itemRequired) {
      html += '<h3>Required Items</h3>';
      html += '<span class="tag item">' + esc(item.itemRequired) + '</span>';
    }

    // Speaking Stone
    if (item.speakingStone) {
      html += '<h3>Speaking Stone Required</h3>';
      let stName = item.speakingStone.replace(/\s*x\d+/i, '').trim();
      html += '<a class="tag item" onclick="showDetail(\'stone\',\'' + escJs(stName) + '\')">' + esc(item.speakingStone) + '</a>';
    }

    // Crafting materials
    let hasMats = Object.keys(item.materials).length || Object.keys(item.wood).length || Object.keys(item.ores).length;
    if (hasMats) {
      html += '<h3>Crafting Materials</h3>';
      html += '<table class="recipe-table"><tr><th>Material</th><th>Qty</th><th>Qty (2 Rep)</th><th>Source</th></tr>';

      const renderMats = (group, category) => {
        Object.entries(group).forEach(([mat, info]) => {
          let source = '';
          if (materialToEnemies[mat]) {
            let names = [...new Set(materialToEnemies[mat].map(e => e.name))];
            source = names.map(n => '<a onclick="showDetail(\'enemy\',\'' + escJs(n) + '\')" style="color:var(--red)">' + esc(n) + '</a>').join(', ');
          }
          if (materialToMarket[mat]) {
            if (source) source += ' | ';
            source += '<span style="color:var(--green)">Market</span>';
          }
          if (DATA.harvestLocations[mat]) {
            if (source) source += ' | ';
            source += '<span style="color:var(--orange)">Nodes: ' + esc(DATA.harvestLocations[mat]) + '</span>';
          }
          html += '<tr><td><a onclick="showDetail(\'material\',\'' + escJs(mat) + '\')">' + esc(mat) + '</a></td>';
          html += '<td>' + esc(info.qty || info) + '</td><td>' + esc(info.rep2 || info['2R'] || '') + '</td>';
          html += '<td style="font-size:0.8rem;">' + (source || '-') + '</td></tr>';
        });
      };

      if (Object.keys(item.materials).length) renderMats(item.materials, 'Monster');
      if (Object.keys(item.wood).length) renderMats(item.wood, 'Wood');
      if (Object.keys(item.ores).length) renderMats(item.ores, 'Ore');
      html += '</table>';
    }

    // Plan Route link
    if (hasMats) {
      html += '<div style="margin-top:14px;"><a class="rp-btn rp-btn-primary" style="text-decoration:none;" onclick="planRouteFor(\'' + escJs(item.name) + '\')">Plan Route</a></div>';
    }

    content.innerHTML = html;

  } else if (type === 'material') {
    let html = '<h2>' + esc(identifier) + '</h2>';
    html += '<div class="card-subtitle">Material / Resource</div>';

    // Dropped by enemies
    if (materialToEnemies[identifier]) {
      html += '<h3>Dropped By Enemies</h3>';
      let uniqueEnemies = {};
      materialToEnemies[identifier].forEach(e => {
        if (!uniqueEnemies[e.name]) uniqueEnemies[e.name] = [];
        uniqueEnemies[e.name].push(e);
      });
      Object.entries(uniqueEnemies).forEach(([name, entries]) => {
        html += '<div style="margin-bottom:8px;padding:8px;background:var(--bg);border-radius:4px;">';
        html += '<a onclick="showDetail(\'enemy\',\'' + escJs(name) + '\')" style="font-weight:600;color:var(--red);">' + esc(name) + '</a>';
        html += '<div style="font-size:0.82rem;color:var(--text2);margin-top:4px;">';
        entries.forEach(e => {
          let locs = Object.values(e.locations).join(', ');
          html += ratingStars(e.rating) + ' ' + (locs ? '<span style="color:var(--green)">' + esc(locs) + '</span>' : 'Location varies') + '<br>';
        });
        html += '</div></div>';
      });
    }

    // Bought at market
    if (materialToMarket[identifier]) {
      let m = materialToMarket[identifier];
      html += '<h3>Market Availability</h3>';
      html += '<table class="recipe-table"><tr><th>Town</th><th>Buy</th><th>Buy (2 Rep)</th><th>Sell</th></tr>';
      Object.entries(m.prices).forEach(([town, p]) => {
        html += '<tr><td>' + esc(town) + '</td><td>' + esc(p.buy || '-') + '</td><td>' + esc(p.buy2Rep || '-') + '</td><td>' + esc(p.sell || '-') + '</td></tr>';
      });
      html += '</table>';
    }

    // Harvest locations
    if (DATA.harvestLocations[identifier]) {
      html += '<h3>Harvesting/Mining Locations</h3>';
      html += '<div style="color:var(--green);font-size:0.9rem;">Nodes: ' + esc(DATA.harvestLocations[identifier]) + '</div>';
    }
    if (DATA.resourceLuxCosts[identifier]) {
      html += '<div style="color:var(--gold);font-size:0.85rem;margin-top:4px;">Lux cost to buy x4: ' + esc(DATA.resourceLuxCosts[identifier]) + '</div>';
    }

    // Used in crafting
    if (materialToCraft[identifier]) {
      html += '<h3>Used to Craft</h3>';
      html += '<div class="tag-list" style="gap:6px;">';
      materialToCraft[identifier].forEach(c => {
        html += '<a class="tag craft" onclick="showDetail(\'craft\',\'' + escJs(c.name) + '\')">' + esc(c.name) + '</a>';
      });
      html += '</div>';
    }

    content.innerHTML = html;

  } else if (type === 'stone') {
    let stone = DATA.speakingStones.find(s => s.name === identifier);
    if (!stone) return;
    let html = '<h2>' + esc(stone.name) + '</h2>';
    html += '<div class="card-subtitle">Speaking Stone</div>';
    html += '<div class="stat-row">';
    html += '<span class="stat"><span class="stat-label">Element</span><span class="stat-value">' + esc(stone.element) + '</span></span>';
    html += '<span class="stat"><span class="stat-label">Color</span><span class="stat-value">' + esc(stone.color) + '</span></span>';
    html += '<span class="stat"><span class="stat-label">Available</span><span class="stat-value">' + esc(stone.available) + '</span></span>';
    html += '</div>';
    html += '<div style="margin:8px 0;padding:8px;background:var(--bg);border-radius:4px;font-size:0.9rem;">' + esc(stone.bonus) + '</div>';
    if (stone.lapidaryExchange) {
      html += '<div style="font-size:0.85rem;color:var(--text2);">Lapidary Exchange: ' + esc(stone.lapidaryExchange) + '</div>';
    }

    // Used in crafting
    if (materialToCraft[identifier]) {
      html += '<h3>Used to Craft</h3>';
      html += '<div class="tag-list" style="gap:6px;">';
      materialToCraft[identifier].forEach(c => {
        html += '<a class="tag craft" onclick="showDetail(\'craft\',\'' + escJs(c.name) + '\')">' + esc(c.name) + '</a>';
      });
      html += '</div>';
    }

    content.innerHTML = html;
  }

  overlay.classList.add('active');
}

function findPrereqChain(itemName) {
  for (const chain of DATA.prereqChains) {
    let idx = chain.findIndex(c => c.name === itemName);
    if (idx >= 0) return chain.slice(0, idx + 1);
  }
  return null;
}

// --- Render functions ---
function renderEnemies(filter, search) {
  const grid = document.getElementById('enemy-grid');
  let grouped = {};
  DATA.enemies.forEach(e => {
    if (filter !== 'all' && e.rating !== filter) return;
    if (search && !matchesSearch(e, search)) return;
    if (!grouped[e.name]) grouped[e.name] = [];
    grouped[e.name].push(e);
  });

  if (!Object.keys(grouped).length) {
    grid.innerHTML = '<div class="empty-msg">No enemies match your filters.</div>';
    return;
  }

  let html = '';
  Object.entries(grouped).forEach(([name, entries]) => {
    html += '<div class="card">';
    html += '<div class="card-title"><a onclick="showDetail(\'enemy\',\'' + escJs(name) + '\')">' + esc(name) + '</a></div>';

    entries.forEach(e => {
      html += '<div style="margin-bottom:6px;">';
      html += '<div style="display:flex;align-items:center;gap:8px;margin-bottom:3px;">' + ratingStars(e.rating);
      html += '<span class="stat attack"><span class="stat-label">ATK</span><span class="stat-value">' + esc(e.attack) + '</span></span>';
      html += '<span class="stat defense"><span class="stat-label">DEF</span><span class="stat-value">' + esc(e.defense) + '</span></span>';
      html += '<span class="stat ap"><span class="stat-label">AP</span><span class="stat-value">' + esc(e.ap) + '</span></span>';
      html += '<span class="stat hp"><span class="stat-label">HP</span><span class="stat-value">' + esc(e.hp) + '</span></span>';
      if (e.lux) html += '<span class="stat lux"><span class="stat-label">Lux</span><span class="stat-value">' + esc(e.lux) + '</span></span>';
      html += '</div>';

      if (e.materialDrops.length) {
        html += '<div class="tag-list">';
        e.materialDrops.forEach(d => {
          let matName = d.includes('(') ? d.substring(0, d.indexOf('(')).trim() : d;
          html += '<a class="tag" onclick="event.stopPropagation();showDetail(\'material\',\'' + escJs(matName) + '\')">' + esc(d) + '</a>';
        });
        html += '</div>';
      }
      if (e.itemDrop) html += '<span class="tag item" style="margin-top:3px;">' + esc(e.itemDrop) + '</span>';
      html += '</div>';
    });

    // Compact locations
    let allLocs = {};
    entries.forEach(e => Object.entries(e.locations).forEach(([ch, loc]) => {
      if (!allLocs[ch]) allLocs[ch] = new Set();
      loc.split(',').forEach(l => allLocs[ch].add(l.trim()));
    }));
    if (Object.keys(allLocs).length) {
      html += '<div class="locations-list">';
      Object.entries(allLocs).forEach(([ch, locs]) => {
        html += '<div><span style="color:var(--text2);font-size:0.78rem;">' + esc(ch) + ':</span> <span>' + esc([...locs].join(', ')) + '</span></div>';
      });
      html += '</div>';
    }

    html += '</div>';
  });
  grid.innerHTML = html;
}

function renderArmorWeapons(typeFilter, ratingFilter, search) {
  const grid = document.getElementById('aw-grid');
  let filtered = DATA.armorWeapons.filter(i => {
    if (typeFilter !== 'all' && i.type !== typeFilter) return false;
    if (ratingFilter !== 'all' && i.rating !== ratingFilter) return false;
    if (search && !matchesSearchCraft(i, search)) return false;
    return true;
  });

  if (!filtered.length) {
    grid.innerHTML = '<div class="empty-msg">No items match your filters.</div>';
    return;
  }

  grid.innerHTML = filtered.map(item => renderCraftCard(item, 'armor-weapon')).join('');
}

function renderAccessories(typeFilter, search) {
  const grid = document.getElementById('acc-grid');
  let filtered = DATA.accessories.filter(i => {
    if (typeFilter !== 'all' && i.type !== typeFilter) return false;
    if (search && !matchesSearchCraft(i, search)) return false;
    return true;
  });

  if (!filtered.length) {
    grid.innerHTML = '<div class="empty-msg">No items match your filters.</div>';
    return;
  }

  grid.innerHTML = filtered.map(item => renderCraftCard(item, 'accessory')).join('');
}

function renderCraftCard(item, type) {
  let html = '<div class="card">';
  html += '<div class="card-title"><a onclick="showDetail(\'craft\',\'' + escJs(item.name) + '\')">' + esc(item.name) + '</a>';
  if (item.rating) html += ' ' + ratingStars(item.rating);
  html += '</div>';

  let sub = [];
  if (item.type) sub.push(item.type);
  if (item.city) sub.push(item.city);
  if (item.limitedTo) sub.push('Limited: ' + item.limitedTo);
  html += '<div class="card-subtitle">' + esc(sub.join(' | ')) + '</div>';

  html += '<div class="stat-row">';
  if (item.statIncrease) html += '<span class="stat"><span class="stat-label">Stat</span><span class="stat-value">' + esc(item.statIncrease) + '</span></span>';
  if (item.stoneSlots) html += '<span class="stat"><span class="stat-label">Slots</span><span class="stat-value">' + esc(item.stoneSlots) + '</span></span>';
  if (item.bonusChip) html += '<span class="stat"><span class="stat-label">Bonus</span><span class="stat-value">' + esc(item.bonusChip) + '</span></span>';
  if (item.craftCost && item.craftCost !== '-') html += '<span class="stat lux"><span class="stat-label">Craft</span><span class="stat-value">' + esc(formatCost(item.craftCost)) + '</span></span>';
  if (item.luxCost && item.luxCost !== '-') html += '<span class="stat lux"><span class="stat-label">Lux</span><span class="stat-value">' + esc(formatCost(item.luxCost)) + '</span></span>';
  html += '</div>';

  if (item.effect) {
    html += '<div style="font-size:0.82rem;color:var(--text);margin:4px 0;">' + esc(item.effect) + '</div>';
  }

  if (item.prerequisite) {
    html += '<div class="section-label">Requires</div>';
    html += '<a class="tag craft" onclick="event.stopPropagation();showDetail(\'craft\',\'' + escJs(item.prerequisite) + '\')">' + esc(item.prerequisite) + '</a>';
  }

  // Show materials compactly
  let mats = [];
  Object.entries(item.materials || {}).forEach(([m, info]) => mats.push(m + ' x' + (info.qty || info)));
  Object.entries(item.wood || {}).forEach(([m, info]) => mats.push(m + ' x' + (info.qty || info)));
  Object.entries(item.ores || {}).forEach(([m, info]) => mats.push(m + ' x' + (info.qty || info)));
  if (item.speakingStone) mats.push(item.speakingStone);
  if (item.itemRequired) mats.push(item.itemRequired);

  if (mats.length) {
    html += '<div class="section-label">Materials</div><div class="tag-list">';
    mats.forEach(m => {
      let baseName = m.replace(/\s*x\d+.*/i, '').replace(/\s*\(.*/,'').trim();
      html += '<a class="tag" onclick="event.stopPropagation();showDetail(\'material\',\'' + escJs(baseName) + '\')">' + esc(m) + '</a>';
    });
    html += '</div>';
  }

  html += '</div>';
  return html;
}

function renderMarket() {
  const container = document.getElementById('market-content');
  const towns = ['Mir', 'Razdor', 'Ryba', 'Silny', 'Strofa', 'Vouno', 'Fort Istra Apothecary'];

  let html = '<table class="market-table"><thead>';
  html += '<tr><th rowspan="2">Item</th><th rowspan="2">Effect</th>';
  towns.forEach(t => html += '<th class="town-header" colspan="3">' + esc(t) + '</th>');
  html += '</tr><tr>';
  towns.forEach(() => html += '<th class="buy">Buy</th><th class="buy">2Rep</th><th class="sell">Sell</th>');
  html += '</tr></thead><tbody>';

  DATA.market.forEach(item => {
    html += '<tr><td><a onclick="showDetail(\'material\',\'' + escJs(item.name) + '\')" style="font-weight:600;">' + esc(item.name) + '</a></td>';
    html += '<td style="font-size:0.8rem;color:var(--text2);max-width:200px;">' + esc(item.effect) + '</td>';
    towns.forEach(t => {
      let p = item.prices[t] || {};
      html += '<td>' + esc(p.buy || '-') + '</td><td>' + esc(p.buy2Rep || '-') + '</td><td>' + esc(p.sell || '-') + '</td>';
    });
    html += '</tr>';
  });

  html += '</tbody></table>';
  container.innerHTML = html;
}

function renderBuildings() {
  const container = document.getElementById('buildings-content');
  let html = '<div class="card-grid">';

  DATA.buildings.forEach(b => {
    html += '<div class="card">';
    html += '<div class="card-title">' + esc(b.name) + '</div>';
    if (b.itemRequired) {
      html += '<div class="card-subtitle">Requires: ' + esc(b.itemRequired) + '</div>';
    }

    let mats = [];
    Object.entries(b.wood).forEach(([m, qty]) => mats.push({ name: m, qty, type: 'Wood' }));
    Object.entries(b.ores).forEach(([m, qty]) => mats.push({ name: m, qty, type: 'Ore' }));

    if (mats.length) {
      html += '<div class="section-label">Resources Needed</div><div class="tag-list">';
      mats.forEach(m => {
        html += '<a class="tag" onclick="showDetail(\'material\',\'' + escJs(m.name) + '\')">' + esc(m.name) + ' x' + esc(m.qty) + '</a>';
      });
      html += '</div>';
    }
    html += '</div>';
  });

  html += '</div>';

  // Harvesting locations
  html += '<h3 style="color:var(--gold);margin-top:20px;margin-bottom:10px;">Harvesting & Mining Locations</h3>';
  html += '<table class="recipe-table"><tr><th>Resource</th><th>Nodes</th><th>Lux Cost (x4)</th></tr>';
  Object.entries(DATA.harvestLocations).forEach(([mat, locs]) => {
    html += '<tr><td><a onclick="showDetail(\'material\',\'' + escJs(mat) + '\')">' + esc(mat) + '</a></td>';
    html += '<td style="color:var(--green)">' + esc(locs) + '</td>';
    html += '<td style="color:var(--gold)">' + esc(DATA.resourceLuxCosts[mat] || '-') + '</td></tr>';
  });
  html += '</table>';

  container.innerHTML = html;
}

function renderStones() {
  const grid = document.getElementById('stones-grid');
  const colorMap = { 'Yellow': '#f0c040', 'Orange': '#ffa726', 'Blue': '#42a5f5', 'Green': '#66bb6a',
    'Red': '#ef5350', 'Purple': '#ab47bc', 'White': '#e0e0e0', 'Any color': '#e0e0e0' };

  grid.innerHTML = DATA.speakingStones.map(s => {
    let clr = colorMap[s.color] || 'var(--text)';
    let html = '<div class="card">';
    html += '<div class="card-title"><a onclick="showDetail(\'stone\',\'' + escJs(s.name) + '\')" style="color:' + clr + '">' + esc(s.name) + '</a></div>';
    html += '<div class="card-subtitle">' + esc(s.element) + ' | ' + esc(s.color) + ' | x' + esc(s.available) + '</div>';
    html += '<div style="padding:6px;background:var(--bg);border-radius:4px;font-size:0.9rem;">' + esc(s.bonus) + '</div>';
    if (s.lapidaryExchange) {
      html += '<div style="font-size:0.82rem;color:var(--text2);margin-top:4px;">Exchange: ' + esc(s.lapidaryExchange) + '</div>';
    }

    // What uses this stone
    if (materialToCraft[s.name]) {
      html += '<div class="section-label">Used In</div><div class="tag-list">';
      materialToCraft[s.name].forEach(c => {
        html += '<a class="tag craft" onclick="event.stopPropagation();showDetail(\'craft\',\'' + escJs(c.name) + '\')">' + esc(c.name) + '</a>';
      });
      html += '</div>';
    }

    html += '</div>';
    return html;
  }).join('');
}

function renderMaterialIndex() {
  const container = document.getElementById('material-index');
  const mats = getAllMaterials();
  container.innerHTML = mats.map(m =>
    '<a class="tag" onclick="showDetail(\'material\',\'' + escJs(m) + '\')" style="font-size:0.9rem;padding:6px 12px;">' + esc(m) + '</a>'
  ).join('');
}

// --- Search ---
function matchesSearch(enemy, q) {
  q = q.toLowerCase();
  if (enemy.name.toLowerCase().includes(q)) return true;
  if (enemy.materialDrops.some(d => d.toLowerCase().includes(q))) return true;
  if (enemy.itemDrop && enemy.itemDrop.toLowerCase().includes(q)) return true;
  if (Object.values(enemy.locations).some(l => l.toLowerCase().includes(q))) return true;
  return false;
}

function matchesSearchCraft(item, q) {
  q = q.toLowerCase();
  if (item.name.toLowerCase().includes(q)) return true;
  if (item.city && item.city.toLowerCase().includes(q)) return true;
  if (item.limitedTo && item.limitedTo.toLowerCase().includes(q)) return true;
  if (item.type && item.type.toLowerCase().includes(q)) return true;
  if (item.effect && item.effect.toLowerCase().includes(q)) return true;
  if (item.prerequisite && item.prerequisite.toLowerCase().includes(q)) return true;
  let allMats = [...Object.keys(item.materials || {}), ...Object.keys(item.wood || {}), ...Object.keys(item.ores || {})];
  if (allMats.some(m => m.toLowerCase().includes(q))) return true;
  return false;
}

// --- Event wiring ---
let currentTab = 'enemies';
let enemyFilter = 'all';
let awTypeFilter = 'all';
let awRatingFilter = 'all';
let accTypeFilter = 'all';
let searchQuery = '';

function switchTab(tab) {
  currentTab = tab;
  document.querySelectorAll('.tab').forEach(t => t.classList.toggle('active', t.dataset.tab === tab));
  document.querySelectorAll('.panel').forEach(p => p.classList.toggle('active', p.id === 'panel-' + tab));
  refreshCurrentTab();
  if (tab === 'route-planner' && typeof rpInit === 'function') rpInit();
}

function refreshCurrentTab() {
  if (currentTab === 'enemies') renderEnemies(enemyFilter, searchQuery);
  else if (currentTab === 'armor-weapons') renderArmorWeapons(awTypeFilter, awRatingFilter, searchQuery);
  else if (currentTab === 'accessories') renderAccessories(accTypeFilter, searchQuery);
}

document.getElementById('tabs').addEventListener('click', e => {
  if (e.target.classList.contains('tab')) switchTab(e.target.dataset.tab);
});

document.getElementById('enemy-filters').addEventListener('click', e => {
  if (e.target.classList.contains('filter-btn')) {
    document.querySelectorAll('#enemy-filters .filter-btn').forEach(b => b.classList.remove('active'));
    e.target.classList.add('active');
    enemyFilter = e.target.dataset.filter;
    renderEnemies(enemyFilter, searchQuery);
  }
});

document.getElementById('aw-filters').addEventListener('click', e => {
  if (e.target.classList.contains('filter-btn')) {
    if (e.target.dataset.filter !== undefined) {
      document.querySelectorAll('#aw-filters .filter-btn[data-filter]').forEach(b => b.classList.remove('active'));
      e.target.classList.add('active');
      awTypeFilter = e.target.dataset.filter;
    }
    if (e.target.dataset.rating !== undefined) {
      document.querySelectorAll('#aw-filters .filter-btn[data-rating]').forEach(b => b.classList.remove('active'));
      e.target.classList.add('active');
      awRatingFilter = e.target.dataset.rating;
    }
    renderArmorWeapons(awTypeFilter, awRatingFilter, searchQuery);
  }
});

document.getElementById('acc-filters').addEventListener('click', e => {
  if (e.target.classList.contains('filter-btn')) {
    document.querySelectorAll('#acc-filters .filter-btn').forEach(b => b.classList.remove('active'));
    e.target.classList.add('active');
    accTypeFilter = e.target.dataset.filter;
    renderAccessories(accTypeFilter, searchQuery);
  }
});

document.getElementById('globalSearch').addEventListener('input', e => {
  searchQuery = e.target.value.trim();
  refreshCurrentTab();
});

document.getElementById('close-detail').addEventListener('click', () => {
  document.getElementById('overlay').classList.remove('active');
});
document.getElementById('overlay').addEventListener('click', e => {
  if (e.target === document.getElementById('overlay')) {
    document.getElementById('overlay').classList.remove('active');
  }
});
document.addEventListener('keydown', e => {
  if (e.key === 'Escape') document.getElementById('overlay').classList.remove('active');
});

// --- Route Planner ---
const RP = {
  initialized: false,
  canvas: null, ctx: null,
  mapImg: null, mapLoaded: false,
  // View state (pan/zoom)
  viewX: 0, viewY: 0, zoom: 1, minZoom: 0.1, maxZoom: 3,
  MAP_W: 3000, MAP_H: 4511,
  // Graph (loaded from DATA or localStorage)
  graph: null,
  adj: {},          // adjacency list: nodeId -> [nodeId, ...]
  distMatrix: null, // lazy-computed BFS distance matrix
  // Route result
  route: null,
  // Editor state
  editMode: false, connectMode: false,
  dragNode: null, dragOffX: 0, dragOffY: 0,
  edgeStart: null,
  // Mouse state
  mouseX: 0, mouseY: 0, isPanning: false, panStartX: 0, panStartY: 0,
  hoverNode: null,
};

function rpValidateGraph(g) {
  if (!g || typeof g !== 'object') return false;
  if (!g.nodes || typeof g.nodes !== 'object') return false;
  if (!Array.isArray(g.edges)) return false;
  for (const [nid, n] of Object.entries(g.nodes)) {
    if (typeof n.x !== 'number' || typeof n.y !== 'number') return false;
    if (n.name !== undefined && typeof n.name !== 'string') return false;
  }
  for (const e of g.edges) {
    if (!Array.isArray(e) || e.length !== 2) return false;
  }
  return true;
}

function rpInit() {
  if (RP.initialized) { rpResize(); rpDraw(); return; }
  RP.initialized = true;
  RP.canvas = document.getElementById('rp-canvas');
  RP.ctx = RP.canvas.getContext('2d');
  // Load graph from localStorage or embedded data
  const saved = localStorage.getItem('tig_map_graph');
  if (saved) {
    try {
      const parsed = JSON.parse(saved);
      if (rpValidateGraph(parsed)) RP.graph = parsed;
    } catch(e) { RP.graph = null; }
  }
  if (!RP.graph) RP.graph = JSON.parse(JSON.stringify(DATA.mapGraph));
  rpBuildAdj();
  // Load map image
  RP.mapImg = new Image();
  RP.mapImg.onload = () => { RP.mapLoaded = true; rpFitView(); rpDraw(); };
  RP.mapImg.src = 'map-of-isofar.png';
  // Populate dropdowns
  rpPopulateItems();
  rpPopulateStartLocations();
  // Event listeners
  rpBindEvents();
  rpResize();
  rpDraw();
}

function rpBuildAdj() {
  RP.adj = {};
  const nodes = RP.graph.nodes;
  for (const nid in nodes) RP.adj[nid] = [];
  RP.graph.edges.forEach(([a, b]) => {
    if (nodes[a] && nodes[b]) {
      if (!RP.adj[a]) RP.adj[a] = [];
      if (!RP.adj[b]) RP.adj[b] = [];
      if (!RP.adj[a].includes(b)) RP.adj[a].push(b);
      if (!RP.adj[b].includes(a)) RP.adj[b].push(a);
    }
  });
  RP.distMatrix = null; // invalidate cache
}

function rpBFS(start) {
  const dist = {}; dist[start] = 0;
  const queue = [start]; let qi = 0;
  while (qi < queue.length) {
    const u = queue[qi++];
    (RP.adj[u] || []).forEach(v => {
      if (dist[v] === undefined) { dist[v] = dist[u] + 1; queue.push(v); }
    });
  }
  return dist;
}

function rpGetDist(a, b) {
  if (!RP.distMatrix) RP.distMatrix = {};
  if (!RP.distMatrix[a]) RP.distMatrix[a] = rpBFS(a);
  const d = RP.distMatrix[a][b];
  return d === undefined ? Infinity : d;
}

function rpBFSPath(start, end) {
  if (start === end) return [start];
  const prev = {}; prev[start] = null;
  const queue = [start]; let qi = 0;
  while (qi < queue.length) {
    const u = queue[qi++];
    for (const v of (RP.adj[u] || [])) {
      if (prev[v] === undefined) {
        prev[v] = u;
        if (v === end) {
          const path = [];
          let c = end;
          while (c !== null) { path.unshift(c); c = prev[c]; }
          return path;
        }
        queue.push(v);
      }
    }
  }
  return null; // unreachable
}

// --- Material Source Resolution ---
function rpGetMaterialSources(materialName, chapter) {
  const sources = [];
  const ch = String(chapter);
  const nodes = RP.graph.nodes;
  // Enemy drops: find enemies that drop this material at nodes available in the selected chapter
  if (materialToEnemies[materialName]) {
    const seen = new Set();
    materialToEnemies[materialName].forEach(enemy => {
      // Only check the exact selected chapter (enemies may not persist across chapters)
      const locStr = enemy.locations['Chapter ' + ch];
      if (locStr) {
        locStr.split(',').forEach(part => {
          let p = part.trim().replace(/\.$/, '');
          // Try as number
          const num = parseInt(p);
          if (!isNaN(num) && nodes[String(num)]) { seen.add(String(num)); return; }
          // Try as special area name
          const saId = rpSpecialAreaId(p);
          if (saId && nodes[saId]) seen.add(saId);
        });
      }
    });
    seen.forEach(nid => sources.push(nid));
  }
  // Market: find towns that sell this material
  if (materialToMarket[materialName]) {
    const m = materialToMarket[materialName];
    Object.keys(m.prices).forEach(town => {
      if (m.prices[town].buy) {
        // Map town name to node ID
        const tid = rpTownId(town);
        if (tid && nodes[tid]) sources.push(tid);
      }
    });
  }
  // Harvest nodes
  if (DATA.harvestLocations[materialName]) {
    const locStr = String(DATA.harvestLocations[materialName]);
    locStr.split(',').forEach(part => {
      const p = part.trim();
      const num = parseInt(p);
      if (!isNaN(num) && nodes[String(num)]) sources.push(String(num));
    });
  }
  return [...new Set(sources)];
}

const _specialAreaMap = {
  "FW - Ice Fields": "fw_ice_fields", "FW - Mount Nebesa": "fw_mount_nebesa",
  "FW - Reka Glacier": "fw_reka_glacier", "FW - Room of Columns": "fw_room_of_columns",
  "FW - Skryvat Temple": "fw_skryvat_temple", "FW - The Broken Lands": "fw_broken_lands",
  "FW - Uchitel Span": "fw_uchitel_span", "FW - Urok Span": "fw_urok_span",
  "FW - Vniz Path": "fw_vniz_path",
  "IC - Abandoned Quarters": "ic_abandoned_quarters", "IC - Abandoned Quartes": "ic_abandoned_quarters",
  "IC - Frozen Lake": "ic_frozen_lake", "IC - Glacial Worm Bones": "ic_glacial_worm_bones",
  "IC - Hall of Ice": "ic_hall_of_ice", "IC - Old Armory": "ic_old_armory", "IC - Ossuary": "ic_ossuary",
};
function rpSpecialAreaId(name) { return _specialAreaMap[name] || null; }

function rpTownId(townName) {
  const map = {
    'Mir': 'mir', 'Razdor': 'razdor', 'Ryba': 'ryba', 'Silny': 'silny',
    'Strofa': 'strofa', 'Vouno': 'vouno',
    'Fort Istra Apothecary': 'fort_istra', 'Fort Istra': 'fort_istra',
  };
  for (const [k, v] of Object.entries(map)) {
    if (townName.includes(k)) return v;
  }
  return null;
}

// --- Multi-Stop Route Optimization ---
function rpComputeRoute(itemName, startNode, chapter) {
  // Determine required materials
  let item = DATA.armorWeapons.find(i => i.name === itemName);
  if (!item) item = DATA.accessories.find(i => i.name === itemName);
  if (!item) item = DATA.buildings.find(i => i.name === itemName);
  if (!item) return null;

  const materialsNeeded = [];
  [item.materials, item.wood, item.ores].forEach(group => {
    Object.entries(group || {}).forEach(([mat, info]) => {
      materialsNeeded.push({ name: mat, qty: info.qty || info });
    });
  });
  if (!materialsNeeded.length) return null;

  // For each material, find source nodes accessible in this chapter
  const matSources = {};
  materialsNeeded.forEach(m => {
    matSources[m.name] = rpGetMaterialSources(m.name, chapter);
  });

  // Check if all materials have at least one source
  const unreachable = materialsNeeded.filter(m => !matSources[m.name].length);
  if (unreachable.length) {
    return { error: 'No accessible sources for: ' + unreachable.map(m => m.name).join(', ') };
  }

  // Greedy nearest-unvisited-source heuristic
  let current = startNode;
  const collected = new Set();
  const stops = []; // { nodeId, materials: [...], dist }

  while (collected.size < materialsNeeded.length) {
    let bestNode = null, bestDist = Infinity, bestMats = [];
    materialsNeeded.forEach(m => {
      if (collected.has(m.name)) return;
      matSources[m.name].forEach(nid => {
        const d = rpGetDist(current, nid);
        if (d < bestDist) {
          bestDist = d;
          bestNode = nid;
          bestMats = [m.name];
        } else if (d === bestDist && nid === bestNode) {
          bestMats.push(m.name);
        }
      });
    });
    if (!bestNode || bestDist === Infinity) {
      return { error: 'Cannot reach all material sources from ' + current };
    }
    // Check if this node provides other uncollected materials too
    const allMatsHere = [];
    materialsNeeded.forEach(m => {
      if (!collected.has(m.name) && matSources[m.name].includes(bestNode)) {
        allMatsHere.push(m.name);
      }
    });
    allMatsHere.forEach(mn => collected.add(mn));
    stops.push({ nodeId: bestNode, materials: allMatsHere, distFromPrev: bestDist });
    current = bestNode;
  }

  // 2-opt improvement
  for (let improved = true; improved;) {
    improved = false;
    for (let i = 0; i < stops.length - 1; i++) {
      for (let j = i + 1; j < stops.length; j++) {
        // Compute current cost of segment i..j
        const before = i === 0 ? startNode : stops[i - 1].nodeId;
        let oldCost = rpGetDist(before, stops[i].nodeId);
        for (let k = i; k < j; k++) oldCost += rpGetDist(stops[k].nodeId, stops[k + 1].nodeId);
        // Compute cost if we reverse i..j
        const rev = stops.slice(i, j + 1).reverse();
        let newCost = rpGetDist(before, rev[0].nodeId);
        for (let k = 0; k < rev.length - 1; k++) newCost += rpGetDist(rev[k].nodeId, rev[k + 1].nodeId);
        if (j + 1 < stops.length) {
          oldCost += rpGetDist(stops[j].nodeId, stops[j + 1].nodeId);
          newCost += rpGetDist(rev[rev.length - 1].nodeId, stops[j + 1].nodeId);
        }
        if (newCost < oldCost) {
          for (let k = 0; k < rev.length; k++) stops[i + k] = rev[k];
          improved = true;
        }
      }
    }
  }

  // Recalculate distances after optimization
  let prev = startNode;
  let totalDist = 0;
  stops.forEach(s => {
    s.distFromPrev = rpGetDist(prev, s.nodeId);
    totalDist += s.distFromPrev;
    prev = s.nodeId;
  });

  // Build full path for drawing
  const fullPath = [];
  prev = startNode;
  stops.forEach(s => {
    const seg = rpBFSPath(prev, s.nodeId);
    if (seg) { if (fullPath.length) seg.shift(); fullPath.push(...seg); }
    prev = s.nodeId;
  });

  return { stops, totalDist, fullPath, startNode, materialsNeeded };
}

// --- Canvas Rendering ---
function rpResize() {
  const area = document.getElementById('rp-map-area');
  if (!area) return;
  RP.canvas.width = area.clientWidth;
  RP.canvas.height = area.clientHeight;
  if (!RP.mapLoaded) rpFitView();
  rpDraw();
}

function rpFitView() {
  if (!RP.canvas) return;
  const cw = RP.canvas.width || 800, ch = RP.canvas.height || 600;
  RP.zoom = Math.min(cw / RP.MAP_W, ch / RP.MAP_H) * 0.95;
  RP.viewX = (cw - RP.MAP_W * RP.zoom) / 2;
  RP.viewY = (ch - RP.MAP_H * RP.zoom) / 2;
}

function rpMapToScreen(mx, my) {
  return [mx * RP.zoom + RP.viewX, my * RP.zoom + RP.viewY];
}
function rpScreenToMap(sx, sy) {
  return [(sx - RP.viewX) / RP.zoom, (sy - RP.viewY) / RP.zoom];
}

function rpDraw() {
  const ctx = RP.ctx;
  if (!ctx) return;
  const cw = RP.canvas.width, ch = RP.canvas.height;
  ctx.clearRect(0, 0, cw, ch);
  ctx.save();
  ctx.translate(RP.viewX, RP.viewY);
  ctx.scale(RP.zoom, RP.zoom);

  // Draw map image
  if (RP.mapLoaded) {
    ctx.drawImage(RP.mapImg, 0, 0, RP.MAP_W, RP.MAP_H);
  } else {
    ctx.fillStyle = '#1a1a2e';
    ctx.fillRect(0, 0, RP.MAP_W, RP.MAP_H);
    ctx.fillStyle = '#666';
    ctx.font = '40px sans-serif';
    ctx.textAlign = 'center';
    ctx.fillText('Loading map...', RP.MAP_W / 2, RP.MAP_H / 2);
  }

  const nodes = RP.graph.nodes;
  const r = Math.max(8, 12 / Math.sqrt(RP.zoom));

  // Draw edges
  ctx.strokeStyle = 'rgba(255,255,255,0.6)';
  ctx.lineWidth = 4 / RP.zoom;
  RP.graph.edges.forEach(([a, b]) => {
    if (!nodes[a] || !nodes[b]) return;
    ctx.beginPath();
    ctx.moveTo(nodes[a].x, nodes[a].y);
    ctx.lineTo(nodes[b].x, nodes[b].y);
    ctx.stroke();
  });

  // Draw connect-mode preview line (dashed line from edgeStart to cursor)
  if (RP.connectMode && RP.edgeStart && nodes[RP.edgeStart]) {
    const sn = nodes[RP.edgeStart];
    const [cmx, cmy] = rpScreenToMap(RP.mouseX, RP.mouseY);
    ctx.save();
    ctx.strokeStyle = '#66bb6a';
    ctx.lineWidth = 3 / RP.zoom;
    ctx.setLineDash([10 / RP.zoom, 8 / RP.zoom]);
    ctx.beginPath();
    ctx.moveTo(sn.x, sn.y);
    ctx.lineTo(cmx, cmy);
    ctx.stroke();
    ctx.setLineDash([]);
    ctx.restore();
  }

  // Draw route path
  if (RP.route && RP.route.fullPath) {
    ctx.strokeStyle = '#e94560';
    ctx.lineWidth = 4 / RP.zoom;
    ctx.lineJoin = 'round';
    ctx.setLineDash([]);
    ctx.beginPath();
    RP.route.fullPath.forEach((nid, i) => {
      const n = nodes[nid];
      if (!n) return;
      if (i === 0) ctx.moveTo(n.x, n.y); else ctx.lineTo(n.x, n.y);
    });
    ctx.stroke();

    // Draw direction arrows along path
    ctx.fillStyle = '#e94560';
    for (let i = 1; i < RP.route.fullPath.length; i++) {
      const prev = nodes[RP.route.fullPath[i - 1]];
      const curr = nodes[RP.route.fullPath[i]];
      if (!prev || !curr) continue;
      const mx = (prev.x + curr.x) / 2, my = (prev.y + curr.y) / 2;
      const angle = Math.atan2(curr.y - prev.y, curr.x - prev.x);
      const sz = 10 / RP.zoom;
      ctx.save();
      ctx.translate(mx, my);
      ctx.rotate(angle);
      ctx.beginPath();
      ctx.moveTo(sz, 0);
      ctx.lineTo(-sz / 2, -sz / 2);
      ctx.lineTo(-sz / 2, sz / 2);
      ctx.closePath();
      ctx.fill();
      ctx.restore();
    }
  }

  // Draw nodes
  for (const [nid, n] of Object.entries(nodes)) {
    let color = '#64b5f6'; // default: numbered node
    if (n.type === 'town') color = '#f0c040';
    else if (n.type === 'special') color = '#ab47bc';
    // Highlight route stops
    let isStop = false, stopIdx = -1;
    if (RP.route && RP.route.stops) {
      const idx = RP.route.stops.findIndex(s => s.nodeId === nid);
      if (idx >= 0) { isStop = true; stopIdx = idx; color = '#ef5350'; }
      if (nid === RP.route.startNode) { isStop = true; stopIdx = -1; color = '#66bb6a'; }
    }
    // Highlight hover
    if (RP.hoverNode === nid) {
      ctx.beginPath();
      ctx.arc(n.x, n.y, r * 1.6, 0, Math.PI * 2);
      ctx.fillStyle = 'rgba(233,69,96,0.3)';
      ctx.fill();
    }
    // Editor: highlight edge start
    if (RP.editMode && RP.edgeStart === nid) {
      ctx.beginPath();
      ctx.arc(n.x, n.y, r * 1.8, 0, Math.PI * 2);
      ctx.strokeStyle = '#66bb6a';
      ctx.lineWidth = 3 / RP.zoom;
      ctx.stroke();
    }
    // Node circle
    ctx.beginPath();
    ctx.arc(n.x, n.y, r, 0, Math.PI * 2);
    ctx.fillStyle = color;
    ctx.fill();
    ctx.strokeStyle = '#000';
    ctx.lineWidth = 1.5 / RP.zoom;
    ctx.stroke();
    // Label
    const fontSize = Math.max(9, 11 / Math.sqrt(RP.zoom));
    ctx.font = 'bold ' + fontSize + 'px sans-serif';
    ctx.textAlign = 'center';
    ctx.textBaseline = 'middle';
    ctx.fillStyle = '#000';
    const label = n.type === 'town' ? n.name : (n.type === 'special' ? '' : nid);
    ctx.fillText(label, n.x, n.y);
    // Route stop number
    if (isStop && stopIdx >= 0) {
      const nr = r * 0.7;
      ctx.beginPath();
      ctx.arc(n.x + r, n.y - r, nr, 0, Math.PI * 2);
      ctx.fillStyle = '#e94560';
      ctx.fill();
      ctx.fillStyle = '#fff';
      ctx.font = 'bold ' + (fontSize * 0.8) + 'px sans-serif';
      ctx.fillText(String(stopIdx + 1), n.x + r, n.y - r);
    }
  }
  ctx.restore();
}

function rpNodeAtScreen(sx, sy) {
  const [mx, my] = rpScreenToMap(sx, sy);
  const hitR = Math.max(15, 20 / RP.zoom);
  let closest = null, closestDist = Infinity;
  for (const [nid, n] of Object.entries(RP.graph.nodes)) {
    const d = Math.hypot(n.x - mx, n.y - my);
    if (d < hitR && d < closestDist) { closest = nid; closestDist = d; }
  }
  return closest;
}

// --- Event Binding ---
function rpBindEvents() {
  const canvas = RP.canvas;
  const area = document.getElementById('rp-map-area');
  const tooltip = document.getElementById('rp-tooltip');

  canvas.addEventListener('mousedown', e => {
    const rect = canvas.getBoundingClientRect();
    const sx = e.clientX - rect.left, sy = e.clientY - rect.top;
    const nid = rpNodeAtScreen(sx, sy);

    // --- Connect mode: click nodes to chain edges ---
    if (RP.editMode && RP.connectMode) {
      if (nid) {
        if (RP.edgeStart && RP.edgeStart !== nid) {
          // Connect previous node to this one
          rpToggleEdge(RP.edgeStart, nid);
          RP.edgeStart = nid; // chain: keep going from this node
        } else if (RP.edgeStart === nid) {
          // Clicked same node: stop connecting
          RP.edgeStart = null;
        } else {
          // First node in chain
          RP.edgeStart = nid;
        }
      } else {
        // Clicked empty space: stop connecting
        RP.edgeStart = null;
      }
      rpDraw();
      return;
    }

    // --- Normal editor mode ---
    if (RP.editMode && nid && e.button === 0) {
      // Start dragging node
      const n = RP.graph.nodes[nid];
      const [nsx, nsy] = rpMapToScreen(n.x, n.y);
      RP.dragNode = nid;
      RP.dragOffX = sx - nsx;
      RP.dragOffY = sy - nsy;
      return;
    }
    if (RP.editMode && !nid && e.button === 0) {
      // Click empty space: add a new node
      const [mx, my] = rpScreenToMap(sx, sy);
      if (mx >= 0 && mx <= RP.MAP_W && my >= 0 && my <= RP.MAP_H) {
        rpAddNodePrompt(Math.round(mx), Math.round(my));
      }
      return;
    }
    // Pan (always available when not in connect mode)
    RP.isPanning = true;
    RP.panStartX = sx - RP.viewX;
    RP.panStartY = sy - RP.viewY;
    canvas.style.cursor = 'grabbing';
  });

  canvas.addEventListener('mousemove', e => {
    const rect = canvas.getBoundingClientRect();
    const sx = e.clientX - rect.left, sy = e.clientY - rect.top;
    RP.mouseX = sx; RP.mouseY = sy;
    if (RP.dragNode) {
      const [mx, my] = rpScreenToMap(sx - RP.dragOffX, sy - RP.dragOffY);
      RP.graph.nodes[RP.dragNode].x = Math.round(mx);
      RP.graph.nodes[RP.dragNode].y = Math.round(my);
      rpDraw();
      return;
    }
    if (RP.isPanning) {
      RP.viewX = sx - RP.panStartX;
      RP.viewY = sy - RP.panStartY;
      rpDraw();
      return;
    }
    // Redraw preview line during connect mode
    if (RP.connectMode && RP.edgeStart) rpDraw();
    // Hover detection
    const nid = rpNodeAtScreen(sx, sy);
    if (nid !== RP.hoverNode) {
      RP.hoverNode = nid;
      rpDraw();
      if (nid) rpShowTooltip(nid, e.clientX, e.clientY);
      else tooltip.style.display = 'none';
    } else if (nid) {
      rpShowTooltip(nid, e.clientX, e.clientY);
    }
  });

  canvas.addEventListener('mouseup', () => {
    RP.dragNode = null;
    RP.isPanning = false;
    canvas.style.cursor = RP.editMode ? 'crosshair' : 'grab';
  });

  canvas.addEventListener('mouseleave', () => {
    RP.dragNode = null;
    RP.isPanning = false;
    RP.hoverNode = null;
    tooltip.style.display = 'none';
    canvas.style.cursor = RP.editMode ? 'crosshair' : 'grab';
    rpDraw();
  });

  canvas.addEventListener('wheel', e => {
    e.preventDefault();
    const rect = canvas.getBoundingClientRect();
    const sx = e.clientX - rect.left, sy = e.clientY - rect.top;
    const [mx, my] = rpScreenToMap(sx, sy);
    const factor = e.deltaY < 0 ? 1.15 : 1 / 1.15;
    const newZoom = Math.min(RP.maxZoom, Math.max(RP.minZoom, RP.zoom * factor));
    RP.viewX = sx - mx * newZoom;
    RP.viewY = sy - my * newZoom;
    RP.zoom = newZoom;
    rpDraw();
  }, { passive: false });

  // Context menu for editor
  canvas.addEventListener('contextmenu', e => {
    e.preventDefault();
    if (!RP.editMode) return;
    const rect = canvas.getBoundingClientRect();
    const sx = e.clientX - rect.left, sy = e.clientY - rect.top;
    const nid = rpNodeAtScreen(sx, sy);
    if (nid) rpEditorContextMenu(nid, e.clientX, e.clientY);
  });

  window.addEventListener('resize', () => { if (RP.initialized) rpResize(); });

  // Calculate button
  document.getElementById('rp-calculate').addEventListener('click', rpCalculateRoute);

  // Editor buttons
  document.getElementById('rp-edit-toggle').addEventListener('click', rpToggleEditor);
  document.getElementById('rp-edit-connect').addEventListener('click', rpToggleConnect);
  document.getElementById('rp-edit-save').addEventListener('click', rpSaveGraph);
  document.getElementById('rp-edit-reset').addEventListener('click', rpResetGraph);
  document.getElementById('rp-edit-export').addEventListener('click', rpExportGraph);

  // Item search filter
  document.getElementById('rp-item-search').addEventListener('input', e => {
    rpFilterItemDropdown(e.target.value);
  });
}

function rpShowTooltip(nid, cx, cy) {
  const tooltip = document.getElementById('rp-tooltip');
  const n = RP.graph.nodes[nid];
  if (!n) return;
  let html = '<div class="tt-name">' + esc(n.name || nid) + '</div>';
  html += '<div class="tt-info">';
  if (n.type) html += 'Type: ' + esc(n.type) + '<br>';
  if (n.chapters && n.chapters.length) html += 'Chapters: ' + n.chapters.map(esc).join(', ') + '<br>';
  if (n.enemies && n.enemies.length) html += 'Enemies: ' + n.enemies.slice(0, 5).map(esc).join(', ') + (n.enemies.length > 5 ? '...' : '') + '<br>';
  if (n.resources && n.resources.length) html += 'Resources: ' + n.resources.map(esc).join(', ') + '<br>';
  html += '</div>';
  tooltip.innerHTML = html;
  tooltip.style.display = 'block';
  const area = document.getElementById('rp-map-area');
  const aRect = area.getBoundingClientRect();
  tooltip.style.left = (cx - aRect.left + 15) + 'px';
  tooltip.style.top = (cy - aRect.top + 15) + 'px';
}

// --- Dropdown Population ---
function rpPopulateItems() {
  const sel = document.getElementById('rp-item-select');
  // Group by type
  const groups = { 'Armor': [], 'Weapon': [], 'Accessory': [], 'Item': [], 'Building': [] };
  DATA.armorWeapons.forEach(i => {
    const t = (i.type || '').includes('Armor') ? 'Armor' : 'Weapon';
    groups[t].push(i.name);
  });
  DATA.accessories.forEach(i => {
    const t = (i.type || '').includes('Accessory') ? 'Accessory' : 'Item';
    groups[t].push(i.name);
  });
  DATA.buildings.forEach(i => {
    groups['Building'].push(i.name);
  });
  let html = '<option value="">-- Select an item --</option>';
  for (const [group, items] of Object.entries(groups)) {
    if (!items.length) continue;
    html += '<optgroup label="' + esc(group) + '">';
    items.sort().forEach(name => {
      html += '<option value="' + esc(name) + '">' + esc(name) + '</option>';
    });
    html += '</optgroup>';
  }
  sel.innerHTML = html;
}

function rpFilterItemDropdown(query) {
  const sel = document.getElementById('rp-item-select');
  const q = query.toLowerCase();
  const prev = sel.value;
  const groups = { 'Armor': [], 'Weapon': [], 'Accessory': [], 'Item': [], 'Building': [] };
  DATA.armorWeapons.forEach(i => {
    const t = (i.type || '').includes('Armor') ? 'Armor' : 'Weapon';
    if (!q || i.name.toLowerCase().includes(q)) groups[t].push(i.name);
  });
  DATA.accessories.forEach(i => {
    const t = (i.type || '').includes('Accessory') ? 'Accessory' : 'Item';
    if (!q || i.name.toLowerCase().includes(q)) groups[t].push(i.name);
  });
  DATA.buildings.forEach(i => {
    if (!q || i.name.toLowerCase().includes(q)) groups['Building'].push(i.name);
  });
  let html = '<option value="">-- Select an item --</option>';
  for (const [group, items] of Object.entries(groups)) {
    if (!items.length) continue;
    html += '<optgroup label="' + esc(group) + '">';
    items.sort().forEach(name => {
      html += '<option value="' + esc(name) + '">' + esc(name) + '</option>';
    });
    html += '</optgroup>';
  }
  sel.innerHTML = html;
  sel.value = prev;
}

function rpPopulateStartLocations() {
  const sel = document.getElementById('rp-start-select');
  let html = '<optgroup label="Towns">';
  for (const [nid, n] of Object.entries(RP.graph.nodes)) {
    if (n.type === 'town') html += '<option value="' + esc(nid) + '">' + esc(n.name) + '</option>';
  }
  html += '</optgroup><optgroup label="Numbered Nodes">';
  const numNodes = Object.entries(RP.graph.nodes)
    .filter(([nid, n]) => !n.type)
    .sort((a, b) => parseInt(a[0]) - parseInt(b[0]));
  numNodes.forEach(([nid, n]) => {
    html += '<option value="' + esc(nid) + '">Node ' + esc(nid) + '</option>';
  });
  html += '</optgroup>';
  sel.innerHTML = html;
}

// --- Route Calculation UI ---
function rpCalculateRoute() {
  const itemName = document.getElementById('rp-item-select').value;
  const startNode = document.getElementById('rp-start-select').value;
  const chapter = document.getElementById('rp-chapter-select').value;
  const resultsDiv = document.getElementById('rp-route-results');
  const summaryDiv = document.getElementById('rp-route-summary');

  if (!itemName) { resultsDiv.innerHTML = '<div class="empty-msg">Please select an item to craft.</div>'; return; }

  const result = rpComputeRoute(itemName, startNode, chapter);
  if (!result) { resultsDiv.innerHTML = '<div class="empty-msg">No crafting materials found for this item.</div>'; summaryDiv.style.display = 'none'; return; }
  if (result.error) { resultsDiv.innerHTML = '<div class="empty-msg" style="color:var(--red);">' + esc(result.error) + '</div>'; summaryDiv.style.display = 'none'; return; }

  RP.route = result;

  // Show summary
  summaryDiv.style.display = 'block';
  summaryDiv.innerHTML = '<span class="label">Total steps:</span> <span class="value">' + result.totalDist + '</span> | <span class="label">Stops:</span> <span class="value">' + result.stops.length + '</span>';

  // Show step-by-step
  let html = '<div class="rp-step" onclick="rpZoomToNode(\'' + esc(startNode) + '\')">';
  html += '<div class="rp-step-num" style="background:var(--green);">S</div>';
  html += '<div class="rp-step-info"><div class="rp-step-node">Start: ' + esc(RP.graph.nodes[startNode].name || startNode) + '</div></div></div>';

  result.stops.forEach((s, i) => {
    const n = RP.graph.nodes[s.nodeId];
    const stepId = 'rp-stop-' + i;
    html += '<div class="rp-step">';
    html += '<div class="rp-step-num" onclick="rpZoomToNode(\'' + esc(s.nodeId) + '\')" style="cursor:pointer;">' + (i + 1) + '</div>';
    html += '<div class="rp-step-info">';
    html += '<div class="rp-step-node" onclick="rpZoomToNode(\'' + esc(s.nodeId) + '\')" style="cursor:pointer;">' + esc(n ? n.name || s.nodeId : s.nodeId) + '</div>';
    html += '<div class="rp-step-mats">Collect: ';
    html += s.materials.map(mat => '<span class="rp-mat-link" onclick="event.stopPropagation();rpToggleMatDetail(\'' + stepId + '\',\'' + escJs(mat) + '\',\'' + escJs(s.nodeId) + '\')">' + esc(mat) + '</span>').join(', ');
    html += '</div>';
    html += '<div class="rp-step-dist">' + s.distFromPrev + ' step' + (s.distFromPrev !== 1 ? 's' : '') + ' from previous</div>';
    html += '<div id="' + esc(stepId) + '-detail"></div>';
    html += '</div></div>';
  });
  resultsDiv.innerHTML = html;
  rpDraw();
}

function rpZoomToNode(nid) {
  const n = RP.graph.nodes[nid];
  if (!n) return;
  const cw = RP.canvas.width, ch = RP.canvas.height;
  RP.zoom = 0.8;
  RP.viewX = cw / 2 - n.x * RP.zoom;
  RP.viewY = ch / 2 - n.y * RP.zoom;
  rpDraw();
}

function rpToggleMatDetail(stepId, matName, nodeId) {
  const el = document.getElementById(stepId + '-detail');
  if (!el) return;
  // Toggle: if already showing this material, close it
  if (el.dataset.showing === matName) { el.innerHTML = ''; el.dataset.showing = ''; return; }
  el.dataset.showing = matName;

  let html = '<div class="rp-mat-detail">';
  html += '<b>' + esc(matName) + '</b> at <b>' + esc(nodeId) + '</b>:<br>';
  let found = false;

  // Enemies that drop this material at this node
  if (materialToEnemies[matName]) {
    const nodeNum = nodeId;
    const nodeNode = RP.graph.nodes[nodeId];
    const nodeName = nodeNode ? nodeNode.name || nodeId : nodeId;
    materialToEnemies[matName].forEach(enemy => {
      // Check if this enemy appears at this node in any chapter
      const chaptersHere = [];
      for (const [ch, locStr] of Object.entries(enemy.locations)) {
        const parts = locStr.split(',').map(p => p.trim().replace(/\.$/, ''));
        // Match numbered nodes or special area IDs
        const matches = parts.some(p => {
          if (p === nodeNum) return true;
          const saId = _specialAreaMap[p];
          if (saId === nodeId) return true;
          return false;
        });
        if (matches) chaptersHere.push(ch.replace('Chapter ', 'Ch'));
      }
      if (chaptersHere.length) {
        found = true;
        html += '<span class="rp-md-enemy">' + esc(enemy.name) + '</span> ' + esc(enemy.rating || '') + ' (ATK ' + esc(enemy.attack) + ' DEF ' + esc(enemy.defense) + ' HP ' + esc(enemy.hp) + ') <span style="color:var(--text2);">' + chaptersHere.join(', ') + '</span><br>';
      }
    });
  }

  // Market sources
  if (materialToMarket[matName]) {
    const m = materialToMarket[matName];
    const tid = rpTownId(RP.graph.nodes[nodeId] ? RP.graph.nodes[nodeId].name || '' : '');
    if (tid === nodeId || RP.graph.nodes[nodeId] && RP.graph.nodes[nodeId].type === 'town') {
      Object.entries(m.prices).forEach(([town, p]) => {
        if (p.buy && rpTownId(town) === nodeId) {
          found = true;
          html += '<span class="rp-md-market">Buy at ' + esc(town) + ': ' + esc(p.buy) + ' silver</span><br>';
        }
      });
    }
  }

  // Harvest
  if (DATA.harvestLocations[matName]) {
    const parts = String(DATA.harvestLocations[matName]).split(',').map(p => p.trim());
    if (parts.includes(nodeId)) {
      found = true;
      html += '<span class="rp-md-source">Harvest/mine here</span><br>';
    }
  }

  if (!found) html += '<span style="color:var(--text2);">Source info not available for this node</span>';
  html += '</div>';
  el.innerHTML = html;
}

// --- Known node IDs from game data (for quick-add suggestions) ---
const KNOWN_NODE_IDS = [
  '1','2','5','6','7','8','10','11','12','13','14','15','17','18','19','21','23','24','25','26','27','28','29','30','31','34','35','36','37','38','39','40','41','42','44','45','46','47','50','52','53','57','58','61','62','63','64','65','66','67','68','71','72','73','76','77','80','82','83','84','85','86','87','88','89','90','91','92','93','97','98','99','103','104','105','108'
];
const KNOWN_TOWNS = [
  {id:'mir',name:'Mir'},{id:'vouno',name:'Vouno'},{id:'razdor',name:'Razdor'},{id:'ryba',name:'Ryba'},{id:'silny',name:'Silny'},{id:'strofa',name:'Strofa'},{id:'fort_istra',name:'Fort Istra'}
];
const KNOWN_SPECIAL = [
  {id:'fw_ice_fields',name:'FW - Ice Fields'},{id:'fw_mount_nebesa',name:'FW - Mount Nebesa'},{id:'fw_reka_glacier',name:'FW - Reka Glacier'},
  {id:'fw_room_of_columns',name:'FW - Room of Columns'},{id:'fw_skryvat_temple',name:'FW - Skryvat Temple'},{id:'fw_broken_lands',name:'FW - The Broken Lands'},
  {id:'fw_uchitel_span',name:'FW - Uchitel Span'},{id:'fw_urok_span',name:'FW - Urok Span'},{id:'fw_vniz_path',name:'FW - Vniz Path'},
  {id:'ic_abandoned_quarters',name:'IC - Abandoned Quarters'},{id:'ic_frozen_lake',name:'IC - Frozen Lake'},{id:'ic_glacial_worm_bones',name:'IC - Glacial Worm Bones'},
  {id:'ic_hall_of_ice',name:'IC - Hall of Ice'},{id:'ic_old_armory',name:'IC - Old Armory'},{id:'ic_ossuary',name:'IC - Ossuary'}
];

// --- Editor Mode ---
function rpToggleEditor() {
  RP.editMode = !RP.editMode;
  RP.connectMode = false;
  RP.edgeStart = null;
  const btn = document.getElementById('rp-edit-toggle');
  const connectBtn = document.getElementById('rp-edit-connect');
  const saveBtn = document.getElementById('rp-edit-save');
  const resetBtn = document.getElementById('rp-edit-reset');
  const exportBtn = document.getElementById('rp-edit-export');
  const helpDiv = document.getElementById('rp-editor-help');
  btn.textContent = RP.editMode ? 'Exit Editor' : 'Edit Map';
  btn.style.background = RP.editMode ? 'var(--accent)' : '';
  btn.style.color = RP.editMode ? '#fff' : '';
  const show = RP.editMode ? '' : 'none';
  connectBtn.style.display = show;
  rpStyleConnectBtn();
  saveBtn.style.display = show;
  resetBtn.style.display = show;
  exportBtn.style.display = show;
  helpDiv.style.display = show;
  RP.canvas.style.cursor = RP.editMode ? 'crosshair' : 'grab';
  RP.canvas.classList.toggle('editor-mode', RP.editMode);
  rpDraw();
}

function rpToggleConnect() {
  RP.connectMode = !RP.connectMode;
  RP.edgeStart = null;
  rpStyleConnectBtn();
  RP.canvas.style.cursor = RP.connectMode ? 'pointer' : 'crosshair';
  rpDraw();
}

function rpStyleConnectBtn() {
  const connectBtn = document.getElementById('rp-edit-connect');
  if (RP.connectMode) {
    connectBtn.textContent = 'Stop Connecting';
    connectBtn.style.background = '#66bb6a';
    connectBtn.style.color = '#000';
  } else {
    connectBtn.textContent = 'Connect Nodes';
    connectBtn.style.background = '';
    connectBtn.style.color = '';
  }
}

function rpAddNodePrompt(mx, my) {
  // Build a list of missing known IDs the user hasn't placed yet
  const existing = new Set(Object.keys(RP.graph.nodes));
  const missingNums = KNOWN_NODE_IDS.filter(id => !existing.has(id));
  const missingTowns = KNOWN_TOWNS.filter(t => !existing.has(t.id));
  const missingSpecial = KNOWN_SPECIAL.filter(s => !existing.has(s.id));

  let suggestion = '';
  if (missingNums.length) suggestion = missingNums[0];
  else if (missingTowns.length) suggestion = missingTowns[0].id;

  let promptMsg = 'Enter node ID to place at this location:';
  if (missingNums.length) promptMsg += '\n\nMissing numbered: ' + missingNums.join(', ');
  if (missingTowns.length) promptMsg += '\n\nMissing towns: ' + missingTowns.map(t => t.id + ' (' + t.name + ')').join(', ');
  if (missingSpecial.length) promptMsg += '\n\nMissing special: ' + missingSpecial.map(s => s.id).join(', ');

  const nodeId = prompt(promptMsg, suggestion);
  if (!nodeId || !nodeId.trim()) return;
  let nid = nodeId.trim();

  // Validate: only allow safe characters (letters, numbers, underscores, hyphens, spaces)
  if (!/^[a-zA-Z0-9_ \-]+$/.test(nid)) { alert('Node ID must contain only letters, numbers, spaces, hyphens, and underscores.'); return; }

  // Normalize input: try to match against known IDs flexibly
  const nidLower = nid.toLowerCase().replace(/[\s\-]+/g, '_');
  // Check if input matches a known town by id or name
  const townMatch = KNOWN_TOWNS.find(t => t.id === nid || t.id === nidLower || t.name.toLowerCase() === nid.toLowerCase());
  if (townMatch) nid = townMatch.id;
  // Check if input matches a known special area by id or name
  const specMatch = KNOWN_SPECIAL.find(s => s.id === nid || s.id === nidLower || s.name.toLowerCase() === nid.toLowerCase());
  if (specMatch) nid = specMatch.id;

  if (RP.graph.nodes[nid]) { alert('Node "' + nid + '" already exists. Drag it to move.'); return; }

  // Determine type and name
  let nodeType = undefined;
  let nodeName = nid;
  if (townMatch) { nodeType = 'town'; nodeName = townMatch.name; }
  if (specMatch) { nodeType = 'special'; nodeName = specMatch.name; }

  const nodeData = { x: mx, y: my, name: nodeName, chapters: [], enemies: [], resources: [] };
  if (nodeType) nodeData.type = nodeType;
  RP.graph.nodes[nid] = nodeData;
  rpBuildAdj();
  rpPopulateStartLocations();
  rpDraw();
}

function rpToggleEdge(a, b) {
  const edges = RP.graph.edges;
  const idx = edges.findIndex(([x, y]) => (x === a && y === b) || (x === b && y === a));
  if (idx >= 0) edges.splice(idx, 1);
  else edges.push([a, b]);
  rpBuildAdj();
  rpDraw();
}

function rpEditorContextMenu(nid, cx, cy) {
  const n = RP.graph.nodes[nid];
  if (confirm('Delete node "' + (n.name || nid) + '"? This removes the node and all its edges.')) {
    delete RP.graph.nodes[nid];
    RP.graph.edges = RP.graph.edges.filter(([a, b]) => a !== nid && b !== nid);
    rpBuildAdj();
    rpPopulateStartLocations();
    rpDraw();
  }
}

function rpSaveGraph() {
  localStorage.setItem('tig_map_graph', JSON.stringify(RP.graph));
  alert('Map graph saved to localStorage! (' + Object.keys(RP.graph.nodes).length + ' nodes, ' + RP.graph.edges.length + ' edges)');
}

function rpResetGraph() {
  if (!confirm('Reset to built-in defaults? This discards all placed nodes and edges.')) return;
  localStorage.removeItem('tig_map_graph');
  RP.graph = JSON.parse(JSON.stringify(DATA.mapGraph));
  rpBuildAdj();
  rpPopulateStartLocations();
  rpDraw();
}

function rpExportGraph() {
  const json = JSON.stringify(RP.graph, null, 2);
  // Copy to clipboard
  navigator.clipboard.writeText(json).then(() => {
    alert('Graph JSON copied to clipboard! (' + Object.keys(RP.graph.nodes).length + ' nodes, ' + RP.graph.edges.length + ' edges)\n\nPaste into MAP_GRAPH in build_app.py to make it permanent.');
  }).catch(() => {
    // Fallback: open in a new window
    const w = window.open('', '_blank');
    const pre = w.document.createElement('pre');
    pre.textContent = json;
    w.document.body.appendChild(pre);
  });
}

// --- planRouteFor: called from craft detail overlay ---
function planRouteFor(itemName) {
  document.getElementById('overlay').classList.remove('active');
  switchTab('route-planner');
  // Set the item dropdown
  const sel = document.getElementById('rp-item-select');
  sel.value = itemName;
  document.getElementById('rp-item-search').value = '';
  rpFilterItemDropdown('');
}

// --- Init ---
buildIndices();
renderEnemies('all', '');
renderMarket();
renderBuildings();
renderStones();
renderMaterialIndex();
</script>
</body>
</html>'''
    return html


def main():
    print("Reading Excel file...")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    print("Parsing enemies...")
    enemies = parse_enemies(wb)
    print(f"  Found {len(enemies)} enemy entries")

    print("Parsing armor & weapons...")
    armor_weapons = parse_armor_weapons(wb)
    print(f"  Found {len(armor_weapons)} armor/weapon entries")

    print("Parsing accessories & items...")
    accessories = parse_accessories(wb)
    print(f"  Found {len(accessories)} accessory/item entries")

    print("Parsing market...")
    market = parse_market(wb)
    print(f"  Found {len(market)} market entries")

    print("Parsing buildings...")
    buildings, harvest_locations, lux_costs = parse_buildings(wb)
    print(f"  Found {len(buildings)} building entries")

    print("Parsing speaking stones...")
    stones = parse_speaking_stones(wb)
    print(f"  Found {len(stones)} speaking stones")

    print("Parsing prerequisite chains...")
    prereqs = parse_prereqs(wb)
    print(f"  Found {len(prereqs)} prerequisite chains")

    data = {
        'enemies': enemies,
        'armorWeapons': armor_weapons,
        'accessories': accessories,
        'market': market,
        'buildings': buildings,
        'harvestLocations': harvest_locations,
        'resourceLuxCosts': lux_costs,
        'speakingStones': stones,
        'prereqChains': prereqs,
    }

    print("Building map graph...")
    map_graph = enrich_map_graph(data)
    data['mapGraph'] = map_graph
    print(f"  {len(map_graph['nodes'])} nodes, {len(map_graph['edges'])} edges")

    print("Generating HTML...")
    html = build_html(data)

    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write(html)

    print(f"\nDone! Open this file in your browser:")
    print(f"  {OUTPUT_FILE}")


if __name__ == '__main__':
    main()
